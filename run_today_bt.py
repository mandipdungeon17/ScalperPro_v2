"""
=============================================================================
SCALPER PRO v2 — Professional Today's Backtest (NIFTY + SENSEX)
=============================================================================
Signal types:
  1. BOX_BREAKOUT   — consolidation box + volume breakout  ← main setup
  2. EMA_PULLBACK   — pullback to 13EMA in strong trend
  3. ST_FLIP        — SuperTrend color flip + volume
  4. VWAP_RECLAIM   — price crosses back above/below VWAP
  5. SR_BOUNCE      — bounce at key daily S/R level

Trend engine: 5EMA / 13EMA / 20EMA + SuperTrend (not just EMA20 block)
=============================================================================
"""

import sys, os, warnings, logging
import numpy as np
import pandas as pd
from datetime import date, datetime
from typing import List, Optional, Tuple, Dict

warnings.filterwarnings("ignore")
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.path.insert(0, os.path.dirname(__file__))

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s  %(levelname)-7s  %(message)s",
                    datefmt="%H:%M:%S")
logging.getLogger("scalper.core.index_levels").setLevel(logging.WARNING)
logging.getLogger("scalper.core.premarket_analysis").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)

TARGET_DATE = date(2026, 4, 2)   # default (overridden by CLI)

# ─── Index config ─────────────────────────────────────────────────────────────
INDEX_PARAMS = {
    "NIFTY":     {"lot_size": 75,  "target_pts": 40,  "sl_pts": 20, "strike_int": 50},
    "BANKNIFTY": {"lot_size": 30,  "target_pts": 100, "sl_pts": 50, "strike_int": 100},
    "SENSEX":    {"lot_size": 20,  "target_pts": 120, "sl_pts": 60, "strike_int": 100},
}

# ─── Strategy constants ───────────────────────────────────────────────────────
EMA_FAST, EMA_MID, EMA_SLOW = 5, 13, 20
ST_PERIOD, ST_MULT           = 10, 3.0
DELTA                        = 0.70    # option delta approximation

# Box breakout
BOX_MIN_BARS   = 5      # minimum bars to form a valid box
BOX_MAX_BARS   = 20     # look back at most this many bars
BOX_MAX_PCT    = 0.012  # box height ≤ 1.2% of price

# Trade management
MAX_HOLD_BARS    = 16      # 4 hours on 15-min
NO_TRADE_BEFORE  = "09:30"
NO_TRADE_AFTER   = "15:00"
MAX_TRADES_DAY   = 4       # max 4 trades per day per index
MAX_SAME_DIR_AT_LEVEL = 1  # don't re-enter same direction at same level
MIN_SCORE        = 3       # need 3+ out of 7 to take a trade

# SL/Target sizing
ATR_SL_MULT   = 1.5   # for non-breakout setups: 1.5x ATR
BOX_SL_BUFFER = 0.002  # for breakout: SL = box_low/high + 0.2% buffer
BOX_RR_MULT   = 2.5   # target = entry ± 2.5x box height


# ═══════════════════════════════════════════════════════════════════════════════
# INDICATOR COMPUTATION
# ═══════════════════════════════════════════════════════════════════════════════

def ema(series: pd.Series, span: int) -> pd.Series:
    return series.ewm(span=span, adjust=False).mean()


def supertrend(df: pd.DataFrame, period=ST_PERIOD, mult=ST_MULT):
    """Returns (st_line, is_bullish) as pd.Series."""
    hi, lo, cl = df["high"], df["low"], df["close"]
    tr = pd.concat([(hi - lo),
                    (hi - cl.shift()).abs(),
                    (lo - cl.shift()).abs()], axis=1).max(axis=1)
    atr = tr.ewm(alpha=1/period, adjust=False).mean()

    bu = (hi + lo) / 2 + mult * atr   # basic upper band
    bl = (hi + lo) / 2 - mult * atr   # basic lower band

    fu = bu.copy(); fl = bl.copy()
    for i in range(1, len(df)):
        fu.iat[i] = bu.iat[i] if (bu.iat[i] < fu.iat[i-1] or cl.iat[i-1] > fu.iat[i-1]) else fu.iat[i-1]
        fl.iat[i] = bl.iat[i] if (bl.iat[i] > fl.iat[i-1] or cl.iat[i-1] < fl.iat[i-1]) else fl.iat[i-1]

    st_line = pd.Series(index=df.index, dtype=float)
    is_bull = pd.Series(index=df.index, dtype=bool)
    for i in range(len(df)):
        if i == 0:
            is_bull.iat[i] = True; st_line.iat[i] = fl.iat[i]
        elif is_bull.iat[i-1] and cl.iat[i] < fl.iat[i]:
            is_bull.iat[i] = False; st_line.iat[i] = fu.iat[i]
        elif not is_bull.iat[i-1] and cl.iat[i] > fu.iat[i]:
            is_bull.iat[i] = True;  st_line.iat[i] = fl.iat[i]
        else:
            is_bull.iat[i] = is_bull.iat[i-1]
            st_line.iat[i] = fl.iat[i] if is_bull.iat[i] else fu.iat[i]
    return st_line, is_bull


def rsi_at(closes: pd.Series, period=14) -> float:
    if len(closes) < period + 1: return 50.0
    d = closes.diff()
    g = d.clip(lower=0).ewm(com=period-1, adjust=False).mean().iloc[-1]
    l = (-d).clip(lower=0).ewm(com=period-1, adjust=False).mean().iloc[-1]
    return 100.0 if l == 0 else 100 - 100 / (1 + g / l)


def bollinger_pct_b(closes: pd.Series, period=20, std_mult=2.0) -> float:
    if len(closes) < period: return 0.5
    mid = closes.rolling(period).mean().iloc[-1]
    std = closes.rolling(period).std().iloc[-1]
    up  = mid + std_mult * std
    lo  = mid - std_mult * std
    return float((closes.iloc[-1] - lo) / (up - lo)) if up != lo else 0.5


def session_vwap(df: pd.DataFrame) -> pd.Series:
    """VWAP that resets each day."""
    df = df.copy()
    if "datetime" not in df.columns:
        df["datetime"] = pd.Series(range(len(df)))
    df["_d"]  = pd.to_datetime(df["datetime"]).dt.date
    df["tp"]  = (df["high"] + df["low"] + df["close"]) / 3
    df["tpv"] = df["tp"] * df["volume"]
    result = []
    for _, grp in df.groupby("_d", sort=False):
        v = grp["tpv"].cumsum() / grp["volume"].cumsum().replace(0, np.nan)
        result.extend(v.tolist())
    return pd.Series(result, index=df.index)


def precompute(df: pd.DataFrame) -> pd.DataFrame:
    """Add all indicator columns to df in one pass."""
    df = df.copy().reset_index(drop=True)
    cl = df["close"]

    df["ema5"]  = ema(cl, EMA_FAST)
    df["ema13"] = ema(cl, EMA_MID)
    df["ema20"] = ema(cl, EMA_SLOW)
    df["st"], df["st_bull"] = supertrend(df)

    df["rsi"] = [rsi_at(cl.iloc[max(0, i-28): i+1]) for i in range(len(df))]
    df["bb_pct_b"] = [bollinger_pct_b(cl.iloc[max(0,i-25): i+1]) for i in range(len(df))]

    vol_avg = df["volume"].rolling(20, min_periods=5).mean()
    df["vol_ratio"] = df["volume"] / vol_avg.replace(0, np.nan)

    try:
        df["vwap"] = session_vwap(df)
    except Exception:
        df["vwap"] = cl

    # Trend state
    def trend_state(row):
        e5, e13, e20, stb = row["ema5"], row["ema13"], row["ema20"], row["st_bull"]
        if e5 > e13 > e20 and stb:  return "STRONG_BULL"
        if e5 < e13 < e20 and not stb: return "STRONG_BEAR"
        if e5 > e13 and (e13 > e20 or stb): return "BULL"
        if e5 < e13 and (e13 < e20 or not stb): return "BEAR"
        return "NEUTRAL"
    df["trend"] = df.apply(trend_state, axis=1)

    return df


# ═══════════════════════════════════════════════════════════════════════════════
# SIGNAL SCORING
# ═══════════════════════════════════════════════════════════════════════════════

def score_bar(row: pd.Series, direction: str) -> Tuple[int, List[str]]:
    """0–7 score. Each indicator adds 1 point."""
    score = 0; reasons = []

    # 1. EMA5 vs EMA13
    if direction == "CE" and row["ema5"] > row["ema13"]:
        score += 1; reasons.append("EMA5>EMA13")
    elif direction == "PE" and row["ema5"] < row["ema13"]:
        score += 1; reasons.append("EMA5<EMA13")

    # 2. EMA13 vs EMA20
    if direction == "CE" and row["ema13"] > row["ema20"]:
        score += 1; reasons.append("EMA13>EMA20")
    elif direction == "PE" and row["ema13"] < row["ema20"]:
        score += 1; reasons.append("EMA13<EMA20")

    # 3. SuperTrend
    stb = bool(row["st_bull"])
    if direction == "CE" and stb:
        score += 1; reasons.append("ST-green")
    elif direction == "PE" and not stb:
        score += 1; reasons.append("ST-red")

    # 4. RSI
    rsi = float(row["rsi"])
    if direction == "CE" and 35 <= rsi <= 68:
        score += 1; reasons.append(f"RSI={rsi:.0f}")
    elif direction == "PE" and 32 <= rsi <= 65:
        score += 1; reasons.append(f"RSI={rsi:.0f}")

    # 5. Bollinger
    pct_b = float(row["bb_pct_b"])
    if direction == "CE" and pct_b <= 0.45:
        score += 1; reasons.append(f"BB={pct_b:.2f}")
    elif direction == "PE" and pct_b >= 0.55:
        score += 1; reasons.append(f"BB={pct_b:.2f}")

    # 6. Volume
    vr = float(row["vol_ratio"])
    if vr >= 1.2:
        score += 1; reasons.append(f"Vol={vr:.1f}x")

    # 7. VWAP
    cl, vwap = float(row["close"]), float(row["vwap"])
    if direction == "CE" and cl > vwap:
        score += 1; reasons.append("AboveVWAP")
    elif direction == "PE" and cl < vwap:
        score += 1; reasons.append("BelowVWAP")

    return score, reasons


# ═══════════════════════════════════════════════════════════════════════════════
# SIGNAL DETECTORS
# ═══════════════════════════════════════════════════════════════════════════════

def detect_box_breakout(df: pd.DataFrame, i: int) -> Optional[dict]:
    """
    Scan past BOX_MIN..BOX_MAX bars for a tight consolidation.
    If current bar breaks above/below the box, return a signal.
    """
    row   = df.iloc[i]
    close = float(row["close"])
    best  = None
    best_score = -1

    for lookback in range(BOX_MIN_BARS, min(BOX_MAX_BARS + 1, i)):
        window = df.iloc[i - lookback: i]   # bars BEFORE current
        bh     = float(window["high"].max())
        bl     = float(window["low"].min())
        rng_pct = (bh - bl) / close

        if rng_pct > BOX_MAX_PCT:
            continue  # too wide — not a box

        direction = None
        if close > bh:
            direction = "CE"
        elif close < bl:
            direction = "PE"

        if direction is None:
            continue

        score, reasons = score_bar(row, direction)
        reasons = [f"Box({lookback}b) {bl:.0f}-{bh:.0f} ({rng_pct:.1%})"] + reasons
        score += 1   # breakout inherently directional bonus

        # Tighter box = more reliable
        if rng_pct < 0.006:
            score += 1; reasons.append("TightBox(<0.6%)")

        score = min(score, 7)
        if score > best_score:
            best_score = score
            best = {
                "type":      "BOX_BREAKOUT",
                "direction": direction,
                "trigger":   bh if direction == "CE" else bl,
                "box_high":  bh,
                "box_low":   bl,
                "box_bars":  lookback,
                "score":     score,
                "reasons":   reasons,
            }

    return best if (best and best["score"] >= MIN_SCORE) else None


def detect_ema_pullback(df: pd.DataFrame, i: int) -> Optional[dict]:
    """Pullback to 13EMA in a strong trend, then resume."""
    if i < 3: return None
    row  = df.iloc[i]
    prev = df.iloc[i - 1]
    trend = str(row["trend"])

    if trend not in ("BULL", "STRONG_BULL", "BEAR", "STRONG_BEAR"):
        return None

    cl, ema13 = float(row["close"]), float(row["ema13"])

    if trend in ("BULL", "STRONG_BULL"):
        touched = float(prev["low"]) <= float(prev["ema13"]) * 1.003
        resumed = cl > ema13
        direction = "CE"
        if float(row["rsi"]) > 65: return None
    else:
        touched = float(prev["high"]) >= float(prev["ema13"]) * 0.997
        resumed = cl < ema13
        direction = "PE"
        if float(row["rsi"]) < 35: return None

    if not (touched and resumed):
        return None

    score, reasons = score_bar(row, direction)
    reasons = [f"EMA13 pullback in {trend}", f"13EMA={ema13:.0f}"] + reasons

    return {
        "type":      "EMA_PULLBACK",
        "direction": direction,
        "trigger":   ema13,
        "box_high":  None, "box_low": None, "box_bars": 0,
        "score":     min(score, 7),
        "reasons":   reasons,
    } if score >= MIN_SCORE else None


def detect_st_flip(df: pd.DataFrame, i: int) -> Optional[dict]:
    """SuperTrend color change — momentum shift."""
    if i < 1: return None
    row  = df.iloc[i]
    prev = df.iloc[i - 1]

    was_bull = bool(prev["st_bull"])
    now_bull = bool(row["st_bull"])
    if was_bull == now_bull: return None

    direction = "CE" if now_bull else "PE"
    if float(row["vol_ratio"]) < 1.0: return None

    score, reasons = score_bar(row, direction)
    reasons = [f"ST flip {'GREEN' if now_bull else 'RED'}"] + reasons
    score += 1  # directional flip bonus

    return {
        "type":      "ST_FLIP",
        "direction": direction,
        "trigger":   float(row["st"]),
        "box_high":  None, "box_low": None, "box_bars": 0,
        "score":     min(score, 7),
        "reasons":   reasons,
    } if score >= MIN_SCORE else None


def detect_vwap_reclaim(df: pd.DataFrame, i: int) -> Optional[dict]:
    """Price crosses back above/below VWAP."""
    if i < 1: return None
    row  = df.iloc[i]
    prev = df.iloc[i - 1]
    cl, vwap = float(row["close"]), float(row["vwap"])
    pcl, pvwap = float(prev["close"]), float(prev["vwap"])

    if pcl < pvwap and cl > vwap:
        direction = "CE"
    elif pcl > pvwap and cl < vwap:
        direction = "PE"
    else:
        return None

    if float(row["vol_ratio"]) < 1.1: return None

    score, reasons = score_bar(row, direction)
    reasons = [f"VWAP reclaim {'above' if direction=='CE' else 'below'} {vwap:.0f}"] + reasons

    return {
        "type":      "VWAP_RECLAIM",
        "direction": direction,
        "trigger":   vwap,
        "box_high":  None, "box_low": None, "box_bars": 0,
        "score":     min(score, 7),
        "reasons":   reasons,
    } if score >= MIN_SCORE else None


def detect_sr_bounce(df: pd.DataFrame, i: int, marker, index: str) -> Optional[dict]:
    """Bounce at key daily S/R level."""
    row = df.iloc[i]
    spot = float(row["close"])

    try:
        prox = marker.check_proximity(spot, index)
    except Exception:
        return None

    if prox.proximity_zone not in ("AT_LEVEL", "APPROACHING"):
        return None
    if prox.nearest_level is None:
        return None

    lv = prox.nearest_level
    direction = prox.direction
    if direction not in ("CE", "PE"): return None
    if lv.touches < 2 or lv.touches > 5: return None

    # Don't fight very strong trends with a bounce
    trend = str(row["trend"])
    if direction == "CE" and trend == "STRONG_BEAR": return None
    if direction == "PE" and trend == "STRONG_BULL": return None

    score, reasons = score_bar(row, direction)
    reasons = [f"SR {lv.level_type.value}@{lv.price:.0f} ({lv.touches}T)", f"{prox.proximity_zone}"] + reasons

    return {
        "type":      "SR_BOUNCE",
        "direction": direction,
        "trigger":   lv.price,
        "box_high":  None, "box_low": None, "box_bars": 0,
        "score":     min(score, 7),
        "reasons":   reasons,
    } if score >= MIN_SCORE else None


def all_signals(df: pd.DataFrame, i: int, marker, index: str) -> List[dict]:
    """Collect all valid signals at bar i, sorted by score descending."""
    sigs = []
    for fn in [detect_box_breakout, detect_ema_pullback,
               detect_st_flip, detect_vwap_reclaim]:
        s = fn(df, i)
        if s: sigs.append(s)
    if marker:
        s = detect_sr_bounce(df, i, marker, index)
        if s: sigs.append(s)

    # De-duplicate same direction: keep highest score
    best: Dict[str, dict] = {}
    for s in sigs:
        d = s["direction"]
        if d not in best or s["score"] > best[d]["score"]:
            best[d] = s
    return sorted(best.values(), key=lambda x: -x["score"])


# ═══════════════════════════════════════════════════════════════════════════════
# BACKTEST ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

def run_today(index: str, params: dict, daily: pd.DataFrame,
              fifteen: pd.DataFrame,
              start_date: date = None, end_date: date = None) -> Tuple[List[dict], int, int]:
    from scalper.core.index_levels import IndexLevelMarker

    sd = start_date or TARGET_DATE
    ed = end_date   or TARGET_DATE

    mask      = (fifteen["datetime"].dt.date >= sd) & \
                (fifteen["datetime"].dt.date <= ed)
    df_period = fifteen[mask].reset_index(drop=True)

    if len(df_period) < 5:
        logger.warning(f"[{index}] No 15-min bars for {TARGET_DATE}")
        return [], 0, 0

    # Pre-compute indicators on FULL history (for context), then filter
    df_full     = precompute(fifteen)
    period_mask = (df_full["datetime"].dt.date >= sd) & \
                  (df_full["datetime"].dt.date <= ed)
    # We use full df for indicator context; extract period bar indices
    today_indices = df_full[period_mask].index.tolist()

    # Build S/R marker from daily history
    marker = None
    try:
        marker = IndexLevelMarker()
        marker.mark_levels(daily_df=daily, index=index)
    except Exception as e:
        logger.warning(f"[{index}] S/R marker failed: {e}")

    lot_size   = params["lot_size"]
    target_pts = params["target_pts"]
    sl_pts     = params["sl_pts"]

    trades        = []
    signal_count  = 0
    open_trade    = None
    daily_trades  = 0
    # Track which (direction, level) combos we've already traded to avoid re-entry
    traded_level_dir: set = set()

    # Compute rolling ATR (14-bar) for the full df_full
    hi = df_full["high"]; lo = df_full["low"]; cl = df_full["close"]
    tr_s = pd.concat([(hi - lo), (hi - cl.shift()).abs(), (lo - cl.shift()).abs()], axis=1).max(axis=1)
    atr_s = tr_s.ewm(alpha=1/14, adjust=False).mean()

    for gi in today_indices:
        if gi < BOX_MIN_BARS + 2:
            continue

        row      = df_full.iloc[gi]
        bar_dt   = pd.to_datetime(row["datetime"])
        bar_time = bar_dt.strftime("%H:%M")

        # Check exit for open trade
        if open_trade:
            open_trade["holding_bars"] += 1
            open_trade = check_exit(open_trade, row, bar_time)
            if open_trade.get("exit_reason"):
                finalize(open_trade, lot_size)
                trades.append(open_trade)
                open_trade = None

        if open_trade is not None:
            continue
        if bar_time < NO_TRADE_BEFORE or bar_time > NO_TRADE_AFTER:
            continue
        if daily_trades >= MAX_TRADES_DAY:
            continue

        # Detect all signals
        sigs = all_signals(df_full, gi, marker, index)
        if not sigs:
            continue

        # BOX_BREAKOUT always takes priority over other signal types
        box_sigs = [s for s in sigs if s["type"] == "BOX_BREAKOUT"]
        if box_sigs:
            sigs = box_sigs  # only consider breakout signals if any exist
        # Else use highest-scoring non-breakout signal

        best_sig = sigs[0]
        signal_count += 1

        # Log all detected signals for transparency
        for s in sigs[:2]:
            logger.info(
                f"[{index}] {bar_time} {s['type']} {s['direction']} "
                f"score={s['score']}/7 | {s['reasons'][:2]}"
            )

        direction = best_sig["direction"]

        # Anti-re-entry: don't take same direction at same approximate level twice
        level_key = (direction, round(float(best_sig["trigger"]) / 100) * 100)
        if level_key in traded_level_dir:
            continue
        traded_level_dir.add(level_key)

        # Entry on signal bar close
        entry_spot = float(row["close"])
        current_atr = float(atr_s.iat[gi]) if not np.isnan(atr_s.iat[gi]) else target_pts

        # SL and Target sizing
        if best_sig["type"] == "BOX_BREAKOUT" and best_sig["box_high"] is not None:
            bh = best_sig["box_high"]; bl = best_sig["box_low"]
            box_h = bh - bl
            buf   = entry_spot * BOX_SL_BUFFER
            if direction == "CE":
                sl_spot     = bl - buf          # SL below box low
                target_spot = entry_spot + max(box_h * BOX_RR_MULT, target_pts)
            else:
                sl_spot     = bh + buf          # SL above box high
                target_spot = entry_spot - max(box_h * BOX_RR_MULT, target_pts)
        else:
            sl_dist = max(current_atr * ATR_SL_MULT, sl_pts)
            if direction == "CE":
                sl_spot     = entry_spot - sl_dist
                target_spot = entry_spot + sl_dist * 2
            else:
                sl_spot     = entry_spot + sl_dist
                target_spot = entry_spot - sl_dist * 2

        daily_trades += 1
        open_trade = {
            "id":            f"{index}-{len(trades)+1:03d}",
            "index":         index,
            "direction":     direction,
            "setup_type":    best_sig["type"],
            "trigger":       best_sig["trigger"],
            "box_high":      best_sig.get("box_high"),
            "box_low":       best_sig.get("box_low"),
            "box_bars":      best_sig.get("box_bars", 0),
            "reasons":       "; ".join(best_sig["reasons"][:4]),
            "entry_date":    bar_dt.strftime("%Y-%m-%d"),
            "entry_time":    bar_time,
            "entry_spot":    entry_spot,
            "target_spot":   target_spot,
            "sl_spot":       sl_spot,
            "score":         best_sig["score"],
            "trend":         str(row["trend"]),
            "rsi":           float(row["rsi"]),
            "bb_pct_b":      float(row["bb_pct_b"]),
            "st_bull":       bool(row["st_bull"]),
            "vol_ratio":     float(row["vol_ratio"]),
            "ema5":          float(row["ema5"]),
            "ema13":         float(row["ema13"]),
            "ema20":         float(row["ema20"]),
            "exit_spot":     None,
            "exit_time":     None,
            "exit_reason":   None,
            "holding_bars":  0,
            "pnl_pts":       0.0,
            "pnl_rs":        0.0,
            "is_winner":     False,
        }

    # EOD force-close
    if open_trade and not open_trade.get("exit_reason"):
        last_gi = today_indices[-1]
        last    = df_full.iloc[last_gi]
        open_trade["exit_spot"]   = float(last["close"])
        open_trade["exit_time"]   = pd.to_datetime(last["datetime"]).strftime("%H:%M")
        open_trade["exit_reason"] = "EOD"
        finalize(open_trade, lot_size)
        trades.append(open_trade)

    return trades, signal_count, len(today_indices)


def check_exit(trade: dict, row: pd.Series, bar_time: str) -> dict:
    if trade.get("exit_reason"):
        return trade

    hi = float(row["high"])
    lo = float(row["low"])
    cl = float(row["close"])
    tm = pd.to_datetime(row["datetime"]).strftime("%H:%M")

    if trade["direction"] == "CE":
        if lo  <= trade["sl_spot"]:
            trade["exit_spot"] = trade["sl_spot"]; trade["exit_reason"] = "SL_HIT"
        elif hi >= trade["target_spot"]:
            trade["exit_spot"] = trade["target_spot"]; trade["exit_reason"] = "TARGET_HIT"
    else:
        if hi  >= trade["sl_spot"]:
            trade["exit_spot"] = trade["sl_spot"]; trade["exit_reason"] = "SL_HIT"
        elif lo <= trade["target_spot"]:
            trade["exit_spot"] = trade["target_spot"]; trade["exit_reason"] = "TARGET_HIT"

    if not trade["exit_reason"] and bar_time >= "15:25":
        trade["exit_spot"] = cl; trade["exit_reason"] = "EOD"
    if not trade["exit_reason"] and trade["holding_bars"] >= MAX_HOLD_BARS:
        trade["exit_spot"] = cl; trade["exit_reason"] = "MAX_HOLD"

    if trade["exit_reason"]:
        trade["exit_time"] = tm

    return trade


def finalize(trade: dict, lot_size: int):
    e = trade["entry_spot"]
    x = trade["exit_spot"] or e
    pts = (x - e) if trade["direction"] == "CE" else (e - x)
    trade["pnl_pts"]   = pts
    trade["pnl_rs"]    = pts * DELTA * lot_size
    trade["is_winner"] = pts > 0


# ═══════════════════════════════════════════════════════════════════════════════
# DATA FETCH
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_data(index: str, fetcher):
    from scalper.data.free_fetcher import FreeDataFetcher
    free = FreeDataFetcher()

    daily = None
    try:
        daily = free.fetch_daily(index, period="2y")
        if daily is not None and "datetime" not in daily.columns:
            daily = daily.reset_index()
            if "Date" in daily.columns:
                daily = daily.rename(columns={"Date": "datetime"})
    except Exception as e:
        logger.warning(f"[{index}] yfinance daily: {e}")

    if daily is None or len(daily) < 60:
        daily = fetcher.fetch_daily_data(index, days_back=500)

    if daily is not None:
        daily["datetime"] = pd.to_datetime(daily["datetime"])

    fifteen = fetcher.fetch_index_data(index, interval="15", days_back=90)
    if fifteen is not None and len(fifteen) >= 50:
        fifteen["datetime"] = pd.to_datetime(fifteen["datetime"])

    return daily, fifteen


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

def export(all_results: list, out_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference

    C_HBG  = "DCE6F1"; C_HDR  = "1F497D"
    C_WIN  = "E2EFDA"; C_LOSS = "FCE4D6"; C_NEU = "FFF2CC"
    thin   = Side(style="thin", color="BFBFBF")
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)
    hf     = lambda: Font(bold=True, color=C_HDR, size=10)
    hfill  = lambda: PatternFill("solid", fgColor=C_HBG)

    def hdr(ws, cols, row=1):
        for ci, c in enumerate(cols, 1):
            cell = ws.cell(row=row, column=ci, value=c)
            cell.font = hf(); cell.fill = hfill()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = brd
        ws.row_dimensions[row].height = 28

    wb = Workbook()

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws = wb.active; ws.title = "Summary"
    ws["A1"] = f"Today Backtest — {TARGET_DATE}   NIFTY + SENSEX"
    ws["A1"].font = Font(bold=True, size=14, color=C_HDR)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:H1"); ws.row_dimensions[1].height = 30

    sum_cols = ["Index", "Trades", "Win", "Loss", "Win%", "Net P&L (Rs)",
                "Profit Factor", "Signals", "Bars"]
    hdr(ws, sum_cols, row=3)

    total_t = total_w = 0; total_pnl = 0.0
    for ri, (index, trades, signals, bars, params) in enumerate(all_results, 4):
        wins   = sum(1 for t in trades if t["is_winner"])
        losses = len(trades) - wins
        gw = sum(t["pnl_rs"] for t in trades if t["is_winner"])
        gl = abs(sum(t["pnl_rs"] for t in trades if not t["is_winner"]))
        net = sum(t["pnl_rs"] for t in trades)
        wp  = wins / len(trades) * 100 if trades else 0
        pf  = (gw / gl) if gl > 0 else (float("inf") if gw > 0 else 0)
        total_t += len(trades); total_w += wins; total_pnl += net

        for ci, v in enumerate([index, len(trades), wins, losses,
                                  f"{wp:.1f}%", round(net, 0),
                                  round(pf, 2) if pf != float("inf") else "inf",
                                  signals, bars], 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = brd; c.alignment = Alignment(horizontal="center")
            if ci == 6:
                c.font = Font(bold=True, color=("375623" if net >= 0 else "9C0006"), size=10)

    tr = len(all_results) + 4
    wr = total_w / total_t * 100 if total_t else 0
    for ci, v in enumerate(["TOTAL", total_t, total_w, total_t - total_w,
                              f"{wr:.1f}%", round(total_pnl, 0), "", "", ""], 1):
        c = ws.cell(row=tr, column=ci, value=v)
        c.border = brd; c.fill = PatternFill("solid", fgColor=C_NEU)
        c.font = Font(bold=True, size=11 if ci == 6 else 10,
                      color=("375623" if total_pnl >= 0 else "9C0006") if ci == 6 else "000000")
        c.alignment = Alignment(horizontal="center")

    for col, w in zip("ABCDEFGHI", [12, 8, 6, 6, 8, 16, 14, 10, 8]):
        ws.column_dimensions[col].width = w

    # ── Per-index trade sheets ────────────────────────────────────────────────
    tcols = ["ID", "Time", "Dir", "Setup", "Trigger", "Box Hi", "Box Lo",
             "Entry", "Target", "SL", "Exit", "Exit Time", "Reason",
             "Hold", "PnL pts", "PnL Rs", "Score", "Trend",
             "RSI", "BB%b", "ST", "Vol Ratio", "Win?", "Reasons"]

    for (index, trades, signals, bars, params) in all_results:
        wt = wb.create_sheet(f"{index}_Today")
        hdr(wt, tcols)

        for r, t in enumerate(trades, 2):
            fill = PatternFill("solid", fgColor=(C_WIN if t["is_winner"] else C_LOSS))
            vals = [
                t["id"], t["entry_time"], t["direction"], t["setup_type"],
                f"{t['trigger']:.0f}" if t["trigger"] else "",
                f"{t['box_high']:.0f}" if t["box_high"] else "",
                f"{t['box_low']:.0f}" if t["box_low"] else "",
                round(t["entry_spot"], 0),
                round(t["target_spot"], 0),
                round(t["sl_spot"], 0),
                round(t["exit_spot"], 0) if t["exit_spot"] else "",
                t["exit_time"] or "",
                t["exit_reason"] or "",
                t["holding_bars"],
                round(t["pnl_pts"], 1), round(t["pnl_rs"], 0),
                t["score"], t["trend"],
                round(t["rsi"], 1), round(t["bb_pct_b"], 2),
                "G" if t["st_bull"] else "R",
                round(t["vol_ratio"], 2),
                "WIN" if t["is_winner"] else "LOSS",
                t["reasons"],
            ]
            for ci, v in enumerate(vals, 1):
                c = wt.cell(row=r, column=ci, value=v)
                c.fill = fill; c.border = brd
                c.alignment = Alignment(horizontal="center" if ci < 24 else "left")

        widths = [10, 7, 5, 14, 9, 9, 9, 9, 9, 9, 9, 9, 12, 6, 8, 10,
                  6, 12, 6, 6, 4, 8, 5, 50]
        for ci, w in enumerate(widths, 1):
            wt.column_dimensions[wt.cell(1, ci).column_letter].width = w

        if len(trades) > 1:
            eq_col = len(tcols) + 2
            wt.cell(1, eq_col, "Equity").font = hf()
            cum = 0
            for ri, t in enumerate(trades, 2):
                cum += t["pnl_rs"]
                wt.cell(ri, eq_col, round(cum, 0))
            chart = LineChart()
            chart.title = f"{index} Equity — {TARGET_DATE}"
            chart.style = 10; chart.width = 22; chart.height = 12
            dr = Reference(wt, min_col=eq_col, min_row=1, max_row=len(trades)+1)
            chart.add_data(dr, titles_from_data=True)
            wt.add_chart(chart, f"A{len(trades) + 4}")

        if not trades:
            wt.cell(3, 1, f"No qualifying signals on {TARGET_DATE}").font = Font(italic=True, color="888888")

    wb.save(out_path)
    logger.info(f"Saved: {out_path}")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def export_multi(all_period_results: list, out_path: str):
    """Multi-period, multi-index Excel export."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference

    C_HBG = "DCE6F1"; C_HDR = "1F497D"
    C_WIN = "E2EFDA"; C_LOSS = "FCE4D6"; C_NEU = "FFF2CC"; C_GREY = "F2F2F2"
    thin  = Side(style="thin", color="BFBFBF")
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    hf    = lambda: Font(bold=True, color=C_HDR, size=10)
    hfill = lambda: PatternFill("solid", fgColor=C_HBG)

    def hdr(ws, cols, row=1):
        for ci, c in enumerate(cols, 1):
            cell = ws.cell(row=row, column=ci, value=c)
            cell.font = hf(); cell.fill = hfill()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = brd
        ws.row_dimensions[row].height = 28

    def pnl_font(v):
        return Font(bold=True, color=("375623" if v >= 0 else "9C0006"), size=10)

    wb = Workbook()

    # ── Master Summary sheet ──────────────────────────────────────────────────
    ws_m = wb.active; ws_m.title = "Summary"
    ws_m["A1"] = "SCALPER PRO v2 — Multi-Period Backtest"
    ws_m["A1"].font = Font(bold=True, size=14, color=C_HDR)
    ws_m["A1"].alignment = Alignment(horizontal="center")
    ws_m.merge_cells("A1:J1"); ws_m.row_dimensions[1].height = 30

    sum_cols = ["Period", "Index", "Trades", "Win", "Loss",
                "Win %", "Net P&L (Rs)", "Profit Factor", "Signals", "Bars"]
    hdr(ws_m, sum_cols, row=3)

    cur_row = 4
    for (label, sd, ed, period_results) in all_period_results:
        # Period header row
        ws_m.cell(cur_row, 1, label).font = Font(bold=True, size=11, color="FFFFFF")
        ws_m.merge_cells(f"A{cur_row}:J{cur_row}")
        for c in range(1, 11):
            ws_m.cell(cur_row, c).fill = PatternFill("solid", fgColor="2E4057")
        cur_row += 1

        period_t = period_w = 0; period_pnl = 0.0
        for (index, trades, signals, bars, params) in period_results:
            wins   = sum(1 for t in trades if t["is_winner"])
            losses = len(trades) - wins
            gw = sum(t["pnl_rs"] for t in trades if t["is_winner"])
            gl = abs(sum(t["pnl_rs"] for t in trades if not t["is_winner"]))
            net = sum(t["pnl_rs"] for t in trades)
            wp  = wins / len(trades) * 100 if trades else 0
            pf  = (gw / gl) if gl > 0 else (float("inf") if gw > 0 else 0.0)
            period_t += len(trades); period_w += wins; period_pnl += net

            row_data = [label, index, len(trades), wins, losses,
                        f"{wp:.1f}%", round(net, 0),
                        f"{pf:.2f}" if pf != float("inf") else "inf",
                        signals, bars]
            for ci, v in enumerate(row_data, 1):
                c = ws_m.cell(cur_row, ci, v)
                c.border = brd; c.alignment = Alignment(horizontal="center")
                if ci == 7:
                    c.font = pnl_font(net)
            cur_row += 1

        # Period subtotal
        period_wr = period_w / period_t * 100 if period_t else 0
        for ci, v in enumerate(["", "SUBTOTAL", period_t, period_w,
                                  period_t - period_w, f"{period_wr:.1f}%",
                                  round(period_pnl, 0), "", "", ""], 1):
            c = ws_m.cell(cur_row, ci, v)
            c.fill = PatternFill("solid", fgColor=C_NEU)
            c.font = Font(bold=True, color=("375623" if period_pnl >= 0 else "9C0006") if ci == 7 else "000000",
                          size=10)
            c.border = brd; c.alignment = Alignment(horizontal="center")
        cur_row += 2

    for col, w in zip("ABCDEFGHIJ", [22, 12, 8, 6, 6, 8, 16, 14, 10, 8]):
        ws_m.column_dimensions[col].width = w

    # ── Per-period trade detail sheets ────────────────────────────────────────
    tcols = ["ID", "Date", "Time", "Dir", "Setup", "Trigger",
             "Box Hi", "Box Lo", "Entry", "Target", "SL",
             "Exit", "Exit Time", "Reason", "Hold",
             "PnL pts", "PnL Rs", "Score", "Trend",
             "RSI", "BB%b", "ST", "Vol", "Win?", "Reasons"]

    for (label, sd, ed, period_results) in all_period_results:
        sheet_name = label[:31].replace("/", "-").replace("–", "-")
        ws = wb.create_sheet(sheet_name)
        hdr(ws, tcols)

        r = 2
        for (index, trades, signals, bars, params) in period_results:
            if not trades:
                ws.cell(r, 1, f"[{index}] No trades").font = Font(italic=True, color="888888")
                r += 1; continue

            # Index sub-header
            ws.cell(r, 1, f"── {index} ──").font = Font(bold=True, color=C_HDR)
            ws.merge_cells(f"A{r}:Y{r}")
            ws.cell(r, 1).fill = PatternFill("solid", fgColor=C_HBG)
            r += 1

            for t in trades:
                fill = PatternFill("solid", fgColor=(C_WIN if t["is_winner"] else C_LOSS))
                entry_date = t.get("entry_date", str(sd))
                vals = [
                    t["id"], entry_date, t["entry_time"],
                    t["direction"], t["setup_type"],
                    f"{t['trigger']:.0f}" if t["trigger"] else "",
                    f"{t['box_high']:.0f}" if t.get("box_high") else "",
                    f"{t['box_low']:.0f}"  if t.get("box_low")  else "",
                    round(t["entry_spot"], 0), round(t["target_spot"], 0),
                    round(t["sl_spot"], 0),
                    round(t["exit_spot"], 0) if t["exit_spot"] else "",
                    t.get("exit_time") or "", t.get("exit_reason") or "",
                    t["holding_bars"],
                    round(t["pnl_pts"], 1), round(t["pnl_rs"], 0),
                    t["score"], t["trend"],
                    round(t["rsi"], 1), round(t["bb_pct_b"], 2),
                    "G" if t["st_bull"] else "R",
                    round(t["vol_ratio"], 2),
                    "WIN" if t["is_winner"] else "LOSS",
                    t["reasons"],
                ]
                for ci, v in enumerate(vals, 1):
                    c = ws.cell(r, ci, v)
                    c.fill = fill; c.border = brd
                    c.alignment = Alignment(horizontal="center" if ci < 25 else "left")
                r += 1
            r += 1  # spacer between indices

        widths = [12, 11, 7, 5, 14, 9, 9, 9, 9, 9, 9, 9, 9, 12,
                  6, 8, 10, 6, 12, 6, 6, 4, 7, 5, 55]
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(1, ci).column_letter].width = w

    wb.save(out_path)
    logger.info(f"Saved: {out_path}")


def _print_period(label: str, all_results: list):
    total_t = total_w = 0; total_pnl = 0.0
    print(f"\n{'─'*72}")
    print(f"  {label}")
    print(f"{'─'*72}")
    for (index, trades, signals, bars, params) in all_results:
        wins    = sum(1 for t in trades if t["is_winner"])
        net     = sum(t["pnl_rs"] for t in trades)
        win_pct = wins / len(trades) * 100 if trades else 0
        total_t += len(trades); total_w += wins; total_pnl += net

        print(f"\n  [{index}] Trades={len(trades)} Win={wins} ({win_pct:.0f}%) "
              f"P&L=Rs {net:+,.0f}  Signals={signals}")
        if not trades:
            print(f"    -> No qualifying setup")
        else:
            print(f"    {'ID':<14} {'Time':<6} {'Dir':<4} {'Setup':<14} "
                  f"{'Entry':>8} {'Exit':>8} {'Reason':<12} {'P&L':>10} Sc")
            print("    " + "-" * 78)
            for t in trades:
                pnl_s = f"Rs {t['pnl_rs']:+,.0f}"
                print(f"    {t['id']:<14} {t['entry_time']:<6} {t['direction']:<4} "
                      f"{t['setup_type']:<14} "
                      f"{t['entry_spot']:>8.0f} {(t['exit_spot'] or 0):>8.0f} "
                      f"{(t['exit_reason'] or ''):<12} {pnl_s:>10} {t['score']}/7")
                print(f"       {t['reasons'][:110]}")

    wr = total_w / total_t * 100 if total_t else 0
    print(f"\n  TOTAL  Trades:{total_t}  Win:{total_w} ({wr:.0f}%)  "
          f"Net P&L: Rs {total_pnl:+,.0f}")


def main():
    from scalper.data.fetcher import DataFetcher
    fetcher = DataFetcher()

    # ── Define periods ────────────────────────────────────────────────────────
    monday     = date(2026, 3, 30)
    week_start = date(2026, 3, 24)
    week_end   = date(2026, 3, 28)

    periods = [
        ("Monday 30-Mar-2026",          monday,     monday),
        ("Last Week 24–28 Mar 2026",    week_start, week_end),
    ]
    indices = list(INDEX_PARAMS.keys())   # NIFTY, BANKNIFTY, SENSEX

    print("\n" + "=" * 72)
    print(f"  SCALPER PRO v2 — Multi-Period Backtest")
    print(f"  Indices : {', '.join(indices)}")
    print(f"  Periods : Monday 30-Mar | Last Week 24-28 Mar")
    print(f"  Setups  : BOX_BREAKOUT | EMA_PULLBACK | ST_FLIP | VWAP_RECLAIM | SR_BOUNCE")
    print("=" * 72)

    # Pre-fetch data for all indices once
    data_cache = {}
    for index in indices:
        print(f"\n[{index}] Fetching data...")
        daily, fifteen = fetch_data(index, fetcher)
        data_cache[index] = (daily, fifteen)
        if daily is not None and fifteen is not None:
            print(f"  Daily={len(daily)}bars  15min={len(fifteen)}bars")
        else:
            print(f"  ERROR: insufficient data")

    # Run each period
    all_period_results = []   # list of (label, sd, ed, all_index_results)
    for label, sd, ed in periods:
        period_results = []
        for index in indices:
            daily, fifteen = data_cache[index]
            params = INDEX_PARAMS[index]
            if daily is None or fifteen is None or len(fifteen) < 50:
                period_results.append((index, [], 0, 0, params))
                continue
            trades, signals, bars = run_today(index, params, daily, fifteen,
                                              start_date=sd, end_date=ed)
            period_results.append((index, trades, signals, bars, params))
        all_period_results.append((label, sd, ed, period_results))
        _print_period(label, period_results)

    # Export to Excel
    out = os.path.join(
        os.path.dirname(__file__),
        f"MultiPeriod_BT_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )
    print(f"\nExporting to: {out}")
    export_multi(all_period_results, out)
    print(f"Done. Excel: {out}\n")


if __name__ == "__main__":
    main()
