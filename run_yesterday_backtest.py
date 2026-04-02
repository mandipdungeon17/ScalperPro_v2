"""
Run NIFTY backtest for 1st April 2026 (yesterday) only.
Exports results to Excel at D:/Trading/ScalperPro_v2/
"""

import sys
import os
import logging
import warnings
from datetime import date, datetime

warnings.filterwarnings("ignore")
sys.path.insert(0, os.path.dirname(__file__))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

TARGET_DATE = date(2026, 4, 1)   # 1st April 2026


def main():
    from scalper.data.fetcher import DataFetcher
    from scalper.backtest.nifty_backtest import NiftyBacktest, analyze_long_term_levels

    print("\n" + "=" * 70)
    print(f"  SCALPER PRO v2 — NIFTY BACKTEST  [{TARGET_DATE}]")
    print("=" * 70)

    fetcher = DataFetcher()
    bt = NiftyBacktest(fetcher)

    print("[1/4] Fetching data (daily + 15-min)...")
    bt.fetch_all_data()

    # Suppress noisy per-bar logs
    logging.getLogger("scalper.core.index_levels").setLevel(logging.WARNING)
    logging.getLogger("scalper.core.premarket_analysis").setLevel(logging.WARNING)

    print(f"[2/4] Running intraday backtest for {TARGET_DATE} ...")
    result = bt._run_intraday(
        label=f"1-Apr-2026",
        start_date=TARGET_DATE,
        end_date=TARGET_DATE,
    )
    result.failure_analysis = bt._analyze_failures(result.trades)

    print("[3/4] Analysing 2-year NIFTY key levels...")
    levels_df, levels_summary = analyze_long_term_levels(fetcher)

    out_path = os.path.join(
        os.path.dirname(__file__),
        f"NIFTY_Yesterday_{TARGET_DATE.strftime('%Y%m%d')}_{datetime.now().strftime('%H%M')}.xlsx"
    )

    print(f"[4/4] Exporting to Excel: {out_path}")
    _export(result, levels_df, levels_summary, out_path)

    _print_summary(result, out_path)
    return result, out_path


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def _export(result, levels_df, levels_summary, out_path):
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference

    wb = Workbook()

    # Colour palette
    C_HEADER  = "1F497D"   # dark blue (text)
    C_HBG     = "DCE6F1"   # light blue header background
    C_WIN     = "E2EFDA"   # green row
    C_LOSS    = "FCE4D6"   # red row
    C_NEUTRAL = "F2F2F2"   # grey row
    C_SUP_BG  = "DAEEF3"   # support level row
    C_RES_BG  = "F2DCDB"   # resistance level row

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_font():
        return Font(bold=True, color=C_HEADER, size=10)

    def hdr_fill():
        return PatternFill("solid", fgColor=C_HBG)

    def write_header(ws, cols):
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.font = hdr_font()
            c.fill = hdr_fill()
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        ws.row_dimensions[1].height = 30

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_s = wb.active
    ws_s.title = "Summary"

    ws_s["A1"] = f"NIFTY Backtest — {TARGET_DATE}"
    ws_s["A1"].font = Font(bold=True, size=14, color="1F497D")
    ws_s["A1"].alignment = Alignment(horizontal="center")
    ws_s.merge_cells("A1:D1")
    ws_s.row_dimensions[1].height = 28

    ws_s["A2"] = "Generated"
    ws_s["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    metrics = [
        ("Date",        result.start_date),
        ("Timeframe",   result.timeframe),
        ("Total Bars",  result.total_bars),
        ("Trades",      result.total_trades),
        ("Winners",     result.winners),
        ("Losers",      result.losers),
        ("Win Rate",    f"{result.win_rate:.1f}%"),
        ("Gross Profit",f"Rs {result.gross_profit:,.0f}"),
        ("Gross Loss",  f"Rs {result.gross_loss:,.0f}"),
        ("Net P&L",     f"Rs {result.total_pnl:+,.0f}"),
        ("Profit Factor",f"{result.profit_factor:.2f}"),
        ("Sharpe Ratio",f"{result.sharpe_ratio:.2f}"),
        ("Max Drawdown",f"Rs {result.max_drawdown:,.0f}"),
        ("Best Trade",  f"Rs {result.best_trade_pnl:+,.0f}"),
        ("Worst Trade", f"Rs {result.worst_trade_pnl:+,.0f}"),
        ("Avg Hold Bars",f"{result.avg_hold_bars:.1f}"),
        ("Max Win Streak", result.max_win_streak),
        ("Max Loss Streak", result.max_loss_streak),
    ]

    for row_idx, (k, v) in enumerate(metrics, 4):
        c_k = ws_s.cell(row=row_idx, column=1, value=k)
        c_v = ws_s.cell(row=row_idx, column=2, value=v)
        c_k.font = Font(bold=True, size=10)
        c_k.fill = PatternFill("solid", fgColor=C_HBG)
        c_k.border = border
        c_v.border = border
        c_v.alignment = Alignment(horizontal="right")
        if k == "Net P&L":
            c_v.font = Font(bold=True, color=("375623" if result.total_pnl >= 0 else "9C0006"), size=11)

    for col, width in [("A", 22), ("B", 18), ("C", 18), ("D", 18)]:
        ws_s.column_dimensions[col].width = width

    # Failure analysis section
    if result.failure_analysis:
        ws_s.cell(row=3, column=3, value="Failure Reason").font = hdr_font()
        ws_s.cell(row=3, column=3).fill = hdr_fill()
        ws_s.cell(row=3, column=4, value="Count").font = hdr_font()
        ws_s.cell(row=3, column=4).fill = hdr_fill()
        fa_sorted = sorted(result.failure_analysis.items(), key=lambda x: -x[1])
        for ri, (reason, cnt) in enumerate(fa_sorted, 4):
            ws_s.cell(row=ri, column=3, value=reason).border = border
            ws_s.cell(row=ri, column=4, value=cnt).border = border

    # ── Sheet 2: Trades ───────────────────────────────────────────────────────
    ws_t = wb.create_sheet("Trades_1Apr2026")

    trade_cols = [
        "ID", "Entry Date", "Entry Time", "Direction", "Level Type",
        "Level Price", "Level TF", "Touches", "Strength", "Proximity",
        "Dist ATR", "Entry Spot", "Target Spot", "SL Spot",
        "Exit Spot", "Exit Time", "Exit Reason", "Hold Bars",
        "Idx PnL pts", "Option PnL pts", "P&L (Rs)", "Win?", "Failure Reason",
    ]
    write_header(ws_t, trade_cols)

    for r, t in enumerate(result.trades, 2):
        row_fill = PatternFill("solid", fgColor=(C_WIN if t.is_winner else C_LOSS))
        vals = [
            t.trade_id, t.entry_date, t.entry_time, t.direction, t.level_type,
            t.level_price, t.level_tf, t.level_touches, round(t.level_strength, 3),
            t.proximity_zone, round(t.distance_atr, 3), t.entry_spot,
            t.target_spot, t.sl_spot, t.exit_spot, t.exit_time, t.exit_reason,
            t.holding_bars, round(t.index_pnl_pts, 1), round(t.option_pnl_pts, 1),
            round(t.pnl_rupees, 0), "WIN" if t.is_winner else "LOSS",
            t.failure_reason,
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws_t.cell(row=r, column=ci, value=v)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    # Column widths
    widths = [8, 12, 10, 10, 12, 12, 10, 8, 9, 12, 9, 12, 12, 10, 10, 10, 12, 9, 12, 14, 11, 6, 25]
    for ci, w in enumerate(widths, 1):
        ws_t.column_dimensions[ws_t.cell(1, ci).column_letter].width = w

    # Equity curve chart (if trades exist)
    if result.equity_curve and len(result.equity_curve) > 1:
        eq_start_col = len(trade_cols) + 2
        ws_t.cell(1, eq_start_col, "Equity (Rs)")
        for ri, val in enumerate(result.equity_curve, 2):
            ws_t.cell(ri, eq_start_col, round(val, 0))
        chart = LineChart()
        chart.title = f"Equity Curve — 1 Apr 2026"
        chart.style = 10
        chart.width = 22
        chart.height = 14
        data_ref = Reference(ws_t, min_col=eq_start_col, min_row=1,
                             max_row=len(result.equity_curve) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        ws_t.add_chart(chart, f"A{len(result.trades) + 4}")

    # ── Sheet 3: Key NIFTY Levels ─────────────────────────────────────────────
    ws_l = wb.create_sheet("NIFTY_Key_Levels")

    if levels_df is not None and len(levels_df) > 0:
        write_header(ws_l, list(levels_df.columns))
        for r, row_data in enumerate(levels_df.itertuples(index=False), 2):
            ltype = str(getattr(row_data, "Type", "")).upper()
            row_fill = PatternFill("solid", fgColor=(C_SUP_BG if ltype == "SUPPORT" else C_RES_BG))
            for ci, v in enumerate(row_data, 1):
                cell = ws_l.cell(row=r, column=ci, value=v)
                cell.fill = row_fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
        for ci in range(1, len(levels_df.columns) + 1):
            ws_l.column_dimensions[ws_l.cell(1, ci).column_letter].width = 16

    if levels_summary is not None and len(levels_summary) > 0:
        start_r = len(levels_df) + 4 if levels_df is not None else 4
        ws_l.cell(start_r, 1, "Summary Zones").font = Font(bold=True, size=11, color="1F497D")
        write_header_at(ws_l, list(levels_summary.columns), start_r + 1, border, hdr_font, hdr_fill)
        for r, row_data in enumerate(levels_summary.itertuples(index=False), start_r + 2):
            for ci, v in enumerate(row_data, 1):
                ws_l.cell(row=r, column=ci, value=v).border = border

    wb.save(out_path)
    logger.info(f"Excel saved: {out_path}")


def write_header_at(ws, cols, row_idx, border, hdr_font_fn, hdr_fill_fn):
    from openpyxl.styles import Alignment
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=row_idx, column=ci, value=col)
        c.font = hdr_font_fn()
        c.fill = hdr_fill_fn()
        c.border = border
        c.alignment = Alignment(horizontal="center")


def _print_summary(result, out_path):
    pnl_sign = "+" if result.total_pnl >= 0 else ""
    print("\n" + "=" * 70)
    print(f"  NIFTY  1-Apr-2026 Backtest Results")
    print("=" * 70)
    print(f"  Trades : {result.total_trades}  |  Win: {result.winners}  Loss: {result.losers}")
    print(f"  Win %  : {result.win_rate:.1f}%")
    print(f"  Net P&L: Rs {pnl_sign}{result.total_pnl:,.0f}")
    print(f"  PF     : {result.profit_factor:.2f}  |  Sharpe: {result.sharpe_ratio:.2f}")
    print(f"  Max DD : Rs {result.max_drawdown:,.0f}")
    if result.total_trades == 0:
        print("\n  No trades triggered — NIFTY was not near any qualifying S/R level.")
    if result.trades:
        print("\n  Trade Log:")
        print(f"  {'#':<4} {'Time':<6} {'Dir':<4} {'Level':>8} {'Entry':>8} {'Exit':>8} {'Reason':<14} {'P&L':>10}")
        print("  " + "-" * 66)
        for t in result.trades:
            pnl_str = f"Rs {t.pnl_rupees:+,.0f}"
            print(f"  {t.trade_id:<4} {t.entry_time:<6} {t.direction:<4} "
                  f"{t.level_price:>8.0f} {t.entry_spot:>8.0f} {t.exit_spot:>8.0f} "
                  f"{t.exit_reason:<14} {pnl_str:>10}")
    print("=" * 70)
    print(f"  Excel: {out_path}")
    print("=" * 70 + "\n")


if __name__ == "__main__":
    main()
