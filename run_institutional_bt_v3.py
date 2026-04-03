"""
=============================================================================
SCALPER PRO v3 — Complete Institutional Backtest (Single File)
=============================================================================
Fixes:
  1. Dhan 90-day limit: chunks intraday requests into 85-day windows
  2. All 5 indices: NIFTY, BANKNIFTY, FINNIFTY, MIDCPNIFTY, SENSEX
  3. Sequential periods: Today → Yesterday → Week → Month → 6 Months
     with prompt between each
  4. Single Excel file with separate sheet per period
  5. Self-learning loss analysis: what went wrong on each losing trade,
     and aggregated patterns to avoid
  6. NEVER uses random sample data — only Dhan API or Yahoo Finance

Run:
  cd ScalperPro_v2
  python run_institutional_bt_v3.py
=============================================================================
"""

import sys, os, warnings, logging, time
import numpy as np
import pandas as pd
from datetime import date, datetime, timedelta
from typing import List, Optional, Tuple, Dict
from collections import defaultdict

warnings.filterwarnings("ignore")
if hasattr(sys.stdout, "reconfigure"):
    try: sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except: pass

# ── Path setup: this file should live in ScalperPro_v2/ root ──────────────────
_this_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _this_dir)

logging.basicConfig(level=logging.WARNING, format="%(asctime)s %(levelname)-7s %(message)s", datefmt="%H:%M:%S")
logger = logging.getLogger("inst_bt_v3")
logger.setLevel(logging.INFO)


# ═══════════════════════════════════════════════════════════════════════════════
# INDEX CONFIG (all 5 indices)
# ═══════════════════════════════════════════════════════════════════════════════

INDEX_PARAMS = {
    "NIFTY":      {"lot_size": 65,  "target_pts": 40,  "sl_pts": 20, "strike_int": 50,  "sec_id": "26000"},
    "BANKNIFTY":  {"lot_size": 30,  "target_pts": 100, "sl_pts": 50, "strike_int": 100, "sec_id": "26009"},
    "FINNIFTY":   {"lot_size": 60,  "target_pts": 40,  "sl_pts": 20, "strike_int": 50,  "sec_id": "26037"},
    "MIDCPNIFTY": {"lot_size": 120,  "target_pts": 30,  "sl_pts": 15, "strike_int": 25,  "sec_id": "26074"},
    "SENSEX":     {"lot_size": 20,  "target_pts": 120, "sl_pts": 60, "strike_int": 100, "sec_id": "1"},
}

# Strategy
EMA_FAST, EMA_MID, EMA_SLOW = 5, 13, 20
ST_PERIOD, ST_MULT = 10, 3.0
DELTA = 0.70
BOX_MIN_BARS, BOX_MAX_BARS, BOX_MAX_PCT = 5, 20, 0.012
MAX_HOLD_BARS = 16
NO_TRADE_BEFORE, NO_TRADE_AFTER = "09:45", "15:00"
MAX_TRADES_DAY = 3
MIN_SCORE = 5
MAX_SCORE = 12
ATR_SL_MULT = 1.5
BOX_SL_BUFFER = 0.002
BOX_RR_MULT = 2.5


# ═══════════════════════════════════════════════════════════════════════════════
# DATA FETCH — Dhan with 90-day chunking + Yahoo fallback (NO random data)
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_dhan_chunked(fetcher, index: str, interval: str, days_back: int) -> Optional[pd.DataFrame]:
    """
    Fetch intraday data from Dhan in 85-day chunks to avoid DH-905 error.
    For daily data, no chunking needed.
    """
    from scalper.config.settings import INDEX_CONFIGS
    config = INDEX_CONFIGS.get(index)
    if not config:
        return None

    if interval == "D":
        return fetcher.fetch_index_data(index, interval="D", days_back=days_back)

    # Chunk intraday requests: 85 days per request (Dhan limit is 90)
    CHUNK_DAYS = 85
    all_dfs = []
    end_dt = datetime.now()
    remaining = days_back

    while remaining > 0:
        chunk = min(remaining, CHUNK_DAYS)
        start_dt = end_dt - timedelta(days=chunk)

        from_str = start_dt.strftime("%Y-%m-%d")
        to_str = end_dt.strftime("%Y-%m-%d")

        df = fetcher.fetch_dhan_historical(
            security_id=config.dhan_security_id,
            exchange_segment="IDX_I",
            instrument="INDEX",
            interval=interval,
            from_date=from_str,
            to_date=to_str,
        )
        if df is not None and len(df) > 0:
            all_dfs.append(df)

        end_dt = start_dt - timedelta(days=1)
        remaining -= chunk
        time.sleep(0.3)  # Rate limit

    if not all_dfs:
        return None

    combined = pd.concat(all_dfs, ignore_index=True)
    combined = combined.drop_duplicates(subset=["datetime"]).sort_values("datetime").reset_index(drop=True)
    return combined


def fetch_yahoo_fallback(index: str, period: str = "6mo", interval: str = "15m") -> Optional[pd.DataFrame]:
    """Yahoo Finance fallback — only for indices it supports."""
    try:
        from scalper.data.free_fetcher import FreeDataFetcher
        free = FreeDataFetcher()
        if interval in ("D", "1d"):
            return free.fetch_daily(index, period=period)
        else:
            yf_interval = {"15": "15m", "5": "5m", "60": "1h", "1": "1m"}.get(interval, interval)
            yf_period = {"30": "1mo", "90": "3mo", "180": "6mo"}.get(str(period), "3mo")
            return free.fetch_intraday(index, interval=yf_interval, period=yf_period)
    except Exception as e:
        logger.warning(f"Yahoo fallback failed for {index}: {e}")
        return None


def fetch_data(index: str, days_back: int = 180):
    """
    Fetch daily + 15-min data. Priority: Dhan (chunked) → Yahoo Finance.
    NEVER falls back to random sample data.
    """
    from scalper.data.fetcher import DataFetcher
    fetcher = DataFetcher()

    # Daily data (no chunking needed)
    daily = fetcher.fetch_daily_data(index, days_back=max(days_back + 200, 500))
    if daily is None or len(daily) < 50:
        logger.info(f"[{index}] Dhan daily failed, trying Yahoo...")
        daily = fetch_yahoo_fallback(index, period="2y", interval="D")

    if daily is not None:
        daily["datetime"] = pd.to_datetime(daily["datetime"])

    # 15-min intraday (chunked for Dhan)
    fifteen = fetch_dhan_chunked(fetcher, index, "15", days_back)
    if fifteen is None or len(fifteen) < 50:
        logger.info(f"[{index}] Dhan 15min failed, trying Yahoo...")
        period_map = {30: "1mo", 90: "3mo", 180: "6mo"}
        fifteen = fetch_yahoo_fallback(index, period=period_map.get(days_back, "3mo"), interval="15m")

    if fifteen is not None:
        fifteen["datetime"] = pd.to_datetime(fifteen["datetime"])

    status = "OK" if (daily is not None and fifteen is not None) else "PARTIAL"
    d_len = len(daily) if daily is not None else 0
    f_len = len(fifteen) if fifteen is not None else 0
    source = "Dhan" if f_len > 0 else "Yahoo"
    logger.info(f"[{index}] Data: daily={d_len} 15min={f_len} source={source} status={status}")

    return daily, fifteen


# ═══════════════════════════════════════════════════════════════════════════════
# INDICATORS (same proven logic)
# ═══════════════════════════════════════════════════════════════════════════════

def ema(s, span): return s.ewm(span=span, adjust=False).mean()

def supertrend(df, period=ST_PERIOD, mult=ST_MULT):
    hi,lo,cl = df["high"],df["low"],df["close"]
    tr = pd.concat([(hi-lo),(hi-cl.shift()).abs(),(lo-cl.shift()).abs()],axis=1).max(axis=1)
    atr = tr.ewm(alpha=1/period,adjust=False).mean()
    bu=(hi+lo)/2+mult*atr; bl=(hi+lo)/2-mult*atr; fu=bu.copy(); fl=bl.copy()
    for i in range(1,len(df)):
        fu.iat[i]=bu.iat[i] if (bu.iat[i]<fu.iat[i-1] or cl.iat[i-1]>fu.iat[i-1]) else fu.iat[i-1]
        fl.iat[i]=bl.iat[i] if (bl.iat[i]>fl.iat[i-1] or cl.iat[i-1]<fl.iat[i-1]) else fl.iat[i-1]
    st=pd.Series(index=df.index,dtype=float); bull=pd.Series(index=df.index,dtype=bool)
    for i in range(len(df)):
        if i==0: bull.iat[i]=True; st.iat[i]=fl.iat[i]
        elif bull.iat[i-1] and cl.iat[i]<fl.iat[i]: bull.iat[i]=False; st.iat[i]=fu.iat[i]
        elif not bull.iat[i-1] and cl.iat[i]>fu.iat[i]: bull.iat[i]=True; st.iat[i]=fl.iat[i]
        else: bull.iat[i]=bull.iat[i-1]; st.iat[i]=fl.iat[i] if bull.iat[i] else fu.iat[i]
    return st, bull

def rsi_at(closes, period=14):
    if len(closes)<period+1: return 50.0
    d=closes.diff(); g=d.clip(lower=0).ewm(com=period-1,adjust=False).mean().iloc[-1]
    l=(-d).clip(lower=0).ewm(com=period-1,adjust=False).mean().iloc[-1]
    return 100.0 if l==0 else 100-100/(1+g/l)

def bollinger_pct_b(closes, period=20, std_mult=2.0):
    if len(closes)<period: return 0.5
    mid=closes.rolling(period).mean().iloc[-1]; std=closes.rolling(period).std().iloc[-1]
    up=mid+std_mult*std; lo=mid-std_mult*std
    return float((closes.iloc[-1]-lo)/(up-lo)) if up!=lo else 0.5

def session_vwap(df):
    df=df.copy()
    if "datetime" not in df.columns: return df["close"]
    df["_d"]=pd.to_datetime(df["datetime"]).dt.date
    df["tp"]=(df["high"]+df["low"]+df["close"])/3; df["tpv"]=df["tp"]*df["volume"]
    result=[]
    for _,grp in df.groupby("_d",sort=False):
        v=grp["tpv"].cumsum()/grp["volume"].cumsum().replace(0,np.nan)
        result.extend(v.tolist())
    return pd.Series(result,index=df.index)

def precompute(df):
    df=df.copy().reset_index(drop=True); cl=df["close"]
    df["ema5"]=ema(cl,EMA_FAST); df["ema13"]=ema(cl,EMA_MID); df["ema20"]=ema(cl,EMA_SLOW)
    df["st"],df["st_bull"]=supertrend(df)
    df["rsi"]=[rsi_at(cl.iloc[max(0,i-28):i+1]) for i in range(len(df))]
    df["bb_pct_b"]=[bollinger_pct_b(cl.iloc[max(0,i-25):i+1]) for i in range(len(df))]
    vol_avg=df["volume"].rolling(20,min_periods=5).mean()
    df["vol_ratio"]=df["volume"]/vol_avg.replace(0,np.nan)
    try: df["vwap"]=session_vwap(df)
    except: df["vwap"]=cl
    def ts(row):
        e5,e13,e20,stb=row["ema5"],row["ema13"],row["ema20"],row["st_bull"]
        if e5>e13>e20 and stb: return "STRONG_BULL"
        if e5<e13<e20 and not stb: return "STRONG_BEAR"
        if e5>e13 and (e13>e20 or stb): return "BULL"
        if e5<e13 and (e13<e20 or not stb): return "BEAR"
        return "NEUTRAL"
    df["trend"]=df.apply(ts,axis=1)
    return df


# ═══════════════════════════════════════════════════════════════════════════════
# SCORING — 12-point institutional
# ═══════════════════════════════════════════════════════════════════════════════

def score_bar(row, direction, df=None, gi=None, ctx=None):
    """12-point scoring with institutional add-ons."""
    try:
        from scalper.indicators.institutional_scoring import enhanced_score_bar
        return enhanced_score_bar(row, direction, df, gi, ctx)
    except ImportError:
        # Fallback to base 7-point if institutional modules not installed
        return _base_score(row, direction)

def _base_score(row, direction):
    """Base 7-point fallback."""
    score=0; reasons=[]
    if direction=="CE" and row["ema5"]>row["ema13"]: score+=1; reasons.append("EMA5>13")
    elif direction=="PE" and row["ema5"]<row["ema13"]: score+=1; reasons.append("EMA5<13")
    if direction=="CE" and row["ema13"]>row["ema20"]: score+=1; reasons.append("EMA13>20")
    elif direction=="PE" and row["ema13"]<row["ema20"]: score+=1; reasons.append("EMA13<20")
    if direction=="CE" and bool(row["st_bull"]): score+=1; reasons.append("ST-G")
    elif direction=="PE" and not bool(row["st_bull"]): score+=1; reasons.append("ST-R")
    rsi=float(row["rsi"])
    if direction=="CE" and 35<=rsi<=68: score+=1; reasons.append(f"RSI={rsi:.0f}")
    elif direction=="PE" and 32<=rsi<=65: score+=1; reasons.append(f"RSI={rsi:.0f}")
    pct_b=float(row["bb_pct_b"])
    if direction=="CE" and pct_b<=0.45: score+=1; reasons.append(f"BB={pct_b:.2f}")
    elif direction=="PE" and pct_b>=0.55: score+=1; reasons.append(f"BB={pct_b:.2f}")
    if float(row["vol_ratio"])>=1.2: score+=1; reasons.append(f"Vol={row['vol_ratio']:.1f}x")
    cl,vwap=float(row["close"]),float(row["vwap"])
    if direction=="CE" and cl>vwap: score+=1; reasons.append("AboveVWAP")
    elif direction=="PE" and cl<vwap: score+=1; reasons.append("BelowVWAP")
    return score, reasons


# ═══════════════════════════════════════════════════════════════════════════════
# SIGNAL DETECTORS
# ═══════════════════════════════════════════════════════════════════════════════

def detect_box_breakout(df, i, ctx=None):
    row=df.iloc[i]; close=float(row["close"]); best=None; bs=-1
    for lb in range(BOX_MIN_BARS,min(BOX_MAX_BARS+1,i)):
        w=df.iloc[i-lb:i]; bh=float(w["high"].max()); bl=float(w["low"].min())
        rp=(bh-bl)/close
        if rp>BOX_MAX_PCT: continue
        d=None
        if close>bh: d="CE"
        elif close<bl: d="PE"
        if d is None: continue
        s,r=score_bar(row,d,df,i,ctx); r=[f"Box({lb}b) {bl:.0f}-{bh:.0f}"]+r; s+=1
        if rp<0.006: s+=1; r.append("TightBox")
        if s>bs: bs=s; best={"type":"BOX_BREAKOUT","direction":d,"trigger":bh if d=="CE" else bl,"box_high":bh,"box_low":bl,"box_bars":lb,"score":min(s,MAX_SCORE),"reasons":r}
    return best if best and best["score"]>=MIN_SCORE else None

def detect_ema_pullback(df, i, ctx=None):
    if i<3: return None
    row=df.iloc[i]; prev=df.iloc[i-1]; trend=str(row["trend"])
    if trend not in ("BULL","STRONG_BULL","BEAR","STRONG_BEAR"): return None
    cl,ema13=float(row["close"]),float(row["ema13"])
    if trend in ("BULL","STRONG_BULL"):
        if not (float(prev["low"])<=float(prev["ema13"])*1.003 and cl>ema13): return None
        d="CE"
        if float(row["rsi"])>65: return None
    else:
        if not (float(prev["high"])>=float(prev["ema13"])*0.997 and cl<ema13): return None
        d="PE"
        if float(row["rsi"])<35: return None
    s,r=score_bar(row,d,df,i,ctx); r=[f"EMA13 pullback {trend}"]+r
    return {"type":"EMA_PULLBACK","direction":d,"trigger":ema13,"box_high":None,"box_low":None,"box_bars":0,"score":min(s,MAX_SCORE),"reasons":r} if s>=MIN_SCORE else None

def detect_st_flip(df, i, ctx=None):
    if i<1: return None
    row=df.iloc[i]; prev=df.iloc[i-1]
    was=bool(prev["st_bull"]); now=bool(row["st_bull"])
    if was==now: return None
    d="CE" if now else "PE"
    if float(row["vol_ratio"])<1.0: return None
    s,r=score_bar(row,d,df,i,ctx); r=[f"ST flip {'GREEN' if now else 'RED'}"]+r; s+=1
    return {"type":"ST_FLIP","direction":d,"trigger":float(row["st"]),"box_high":None,"box_low":None,"box_bars":0,"score":min(s,MAX_SCORE),"reasons":r} if s>=MIN_SCORE else None

def detect_vwap_reclaim(df, i, ctx=None):
    if i<1: return None
    row=df.iloc[i]; prev=df.iloc[i-1]
    cl,vwap=float(row["close"]),float(row["vwap"]); pcl,pvwap=float(prev["close"]),float(prev["vwap"])
    if pcl<pvwap and cl>vwap: d="CE"
    elif pcl>pvwap and cl<vwap: d="PE"
    else: return None
    if float(row["vol_ratio"])<1.1: return None
    s,r=score_bar(row,d,df,i,ctx); r=[f"VWAP reclaim {vwap:.0f}"]+r
    return {"type":"VWAP_RECLAIM","direction":d,"trigger":vwap,"box_high":None,"box_low":None,"box_bars":0,"score":min(s,MAX_SCORE),"reasons":r} if s>=MIN_SCORE else None

def detect_sr_bounce(df, i, marker, index, ctx=None):
    row=df.iloc[i]; spot=float(row["close"])
    try: prox=marker.check_proximity(spot,index)
    except: return None
    if prox.proximity_zone not in ("AT_LEVEL","APPROACHING"): return None
    if prox.nearest_level is None: return None
    lv=prox.nearest_level; d=prox.direction
    if d not in ("CE","PE"): return None
    if lv.touches<2: return None
    trend=str(row["trend"])
    if d=="CE" and trend=="STRONG_BEAR": return None
    if d=="PE" and trend=="STRONG_BULL": return None
    s,r=score_bar(row,d,df,i,ctx); r=[f"SR {lv.level_type.value}@{lv.price:.0f} ({lv.touches}T)"]+r
    return {"type":"SR_BOUNCE","direction":d,"trigger":lv.price,"box_high":None,"box_low":None,"box_bars":0,"score":min(s,MAX_SCORE),"reasons":r} if s>=MIN_SCORE else None

def detect_liquidity_sweep(df, i, ctx):
    if ctx is None or not ctx.initialized or not ctx.stop_clusters: return None
    try:
        from scalper.core.liquidity_engine import LiquidityEngine
        sweep = LiquidityEngine().check_current_bar(df, i, ctx.stop_clusters)
        if sweep is None or sweep.score < 5: return None
        d=sweep.trade_direction
        s,r=score_bar(df.iloc[i],d,df,i,ctx); r=[f"LIQ_SWEEP {sweep.sweep_type} @{sweep.cluster_price:.0f}"]+r; s+=2
        return {"type":"LIQ_SWEEP","direction":d,"trigger":sweep.cluster_price,"box_high":None,"box_low":None,"box_bars":0,
                "score":min(s,MAX_SCORE),"reasons":r,"sweep_sl":sweep.sl_zone,"sweep_target":sweep.target_zone} if s>=MIN_SCORE else None
    except: return None

def all_signals(df, i, marker, index, ctx=None):
    sigs=[]
    for fn in [detect_box_breakout,detect_ema_pullback,detect_st_flip,detect_vwap_reclaim]:
        s=fn(df,i,ctx)
        if s: sigs.append(s)
    if marker:
        s=detect_sr_bounce(df,i,marker,index,ctx)
        if s: sigs.append(s)
    if ctx:
        s=detect_liquidity_sweep(df,i,ctx)
        if s: sigs.append(s)
    best={}
    for s in sigs:
        d=s["direction"]
        if d not in best or s["score"]>best[d]["score"]: best[d]=s
    return sorted(best.values(), key=lambda x:-x["score"])


# ═══════════════════════════════════════════════════════════════════════════════
# BACKTEST ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

def check_exit(t, row, bt):
    if t.get("exit_reason"): return t
    hi,lo,cl=float(row["high"]),float(row["low"]),float(row["close"])
    if t["direction"]=="CE":
        if lo<=t["sl_spot"]: t["exit_spot"]=t["sl_spot"]; t["exit_reason"]="SL_HIT"
        elif hi>=t["target_spot"]: t["exit_spot"]=t["target_spot"]; t["exit_reason"]="TARGET_HIT"
    else:
        if hi>=t["sl_spot"]: t["exit_spot"]=t["sl_spot"]; t["exit_reason"]="SL_HIT"
        elif lo<=t["target_spot"]: t["exit_spot"]=t["target_spot"]; t["exit_reason"]="TARGET_HIT"
    if not t["exit_reason"] and bt>="15:25": t["exit_spot"]=cl; t["exit_reason"]="EOD"
    if not t["exit_reason"] and t["holding_bars"]>=MAX_HOLD_BARS: t["exit_spot"]=cl; t["exit_reason"]="MAX_HOLD"
    if t["exit_reason"]: t["exit_time"]=bt
    return t

def finalize(t, lot_size):
    e=t["entry_spot"]; x=t["exit_spot"] or e
    pts=(x-e) if t["direction"]=="CE" else (e-x)
    t["pnl_pts"]=pts; t["pnl_rs"]=pts*DELTA*lot_size; t["is_winner"]=pts>0

def run_period(index, params, daily, fifteen, sd, ed):
    from scalper.core.index_levels import IndexLevelMarker
    mask=(fifteen["datetime"].dt.date>=sd)&(fifteen["datetime"].dt.date<=ed)
    if mask.sum()<5: return [],0,0
    df_full=precompute(fifteen); pm=(df_full["datetime"].dt.date>=sd)&(df_full["datetime"].dt.date<=ed)
    tidx=df_full[pm].index.tolist()
    if not tidx: return [],0,0
    marker=None
    try: marker=IndexLevelMarker(); marker.mark_levels(daily_df=daily,index=index)
    except: pass
    ctx=None
    try:
        from scalper.indicators.institutional_scoring import setup_institutional_context
        ctx=setup_institutional_context(df_full, daily)
    except: pass
    lot=params["lot_size"]; tp=params["target_pts"]; sl=params["sl_pts"]
    trades=[]; sig_count=0; ot=None; dtc={}; tld=set()
    hi=df_full["high"];lo=df_full["low"];cl=df_full["close"]
    tr_s=pd.concat([(hi-lo),(hi-cl.shift()).abs(),(lo-cl.shift()).abs()],axis=1).max(axis=1)
    atr_s=tr_s.ewm(alpha=1/14,adjust=False).mean()
    for gi in tidx:
        if gi<BOX_MIN_BARS+2: continue
        row=df_full.iloc[gi]; bdt=pd.to_datetime(row["datetime"]); bt=bdt.strftime("%H:%M"); bd=bdt.date()
        if bd not in dtc: dtc[bd]=0; tld=set()
        if ot:
            ot["holding_bars"]+=1; ot=check_exit(ot,row,bt)
            if ot.get("exit_reason"): finalize(ot,lot); trades.append(ot); ot=None
        if ot is not None: continue
        if bt<NO_TRADE_BEFORE or bt>NO_TRADE_AFTER: continue
        if dtc[bd]>=MAX_TRADES_DAY: continue
        sigs=all_signals(df_full,gi,marker,index,ctx)
        if not sigs: continue
        liq=[s for s in sigs if s["type"]=="LIQ_SWEEP"]
        box=[s for s in sigs if s["type"]=="BOX_BREAKOUT"]
        if liq: sigs=liq
        elif box: sigs=box
        best=sigs[0]; sig_count+=1; d=best["direction"]
        lk=(d,round(float(best["trigger"])/100)*100)
        if lk in tld: continue
        tld.add(lk)
        es=float(row["close"]); ca=float(atr_s.iat[gi]) if not np.isnan(atr_s.iat[gi]) else tp
        if best["type"]=="BOX_BREAKOUT" and best.get("box_high"):
            bh,bl=best["box_high"],best["box_low"]; bxh=bh-bl; buf=es*BOX_SL_BUFFER
            if d=="CE": sls=bl-buf; tgt=es+max(bxh*BOX_RR_MULT,tp)
            else: sls=bh+buf; tgt=es-max(bxh*BOX_RR_MULT,tp)
        elif best.get("sweep_sl"):
            sls=best["sweep_sl"]; tgt=best["sweep_target"]
        else:
            sd2=max(ca*ATR_SL_MULT,sl)
            if d=="CE": sls=es-sd2; tgt=es+sd2*2
            else: sls=es+sd2; tgt=es-sd2*2
        dtc[bd]+=1
        ot={"id":f"{index}-{len(trades)+1:03d}","index":index,"direction":d,"setup_type":best["type"],
            "trigger":best["trigger"],"box_high":best.get("box_high"),"box_low":best.get("box_low"),
            "box_bars":best.get("box_bars",0),"reasons":"; ".join(best["reasons"][:5]),
            "entry_date":bdt.strftime("%Y-%m-%d"),"entry_time":bt,"entry_spot":es,
            "target_spot":tgt,"sl_spot":sls,"score":best["score"],"max_score":MAX_SCORE,
            "trend":str(row["trend"]),"rsi":float(row["rsi"]),"bb_pct_b":float(row["bb_pct_b"]),
            "st_bull":bool(row["st_bull"]),"vol_ratio":float(row["vol_ratio"]),
            "exit_spot":None,"exit_time":None,"exit_reason":None,"holding_bars":0,
            "pnl_pts":0.0,"pnl_rs":0.0,"is_winner":False}
    if ot and not ot.get("exit_reason"):
        last=df_full.iloc[tidx[-1]]
        ot["exit_spot"]=float(last["close"]); ot["exit_time"]=pd.to_datetime(last["datetime"]).strftime("%H:%M")
        ot["exit_reason"]="EOD"; finalize(ot,lot); trades.append(ot)
    return trades, sig_count, len(tidx)


# ═══════════════════════════════════════════════════════════════════════════════
# SELF-LEARNING LOSS ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════════

def analyze_losses(all_trades: List[dict]) -> List[dict]:
    """Analyze each losing trade to identify what went wrong and patterns to avoid."""
    losers = [t for t in all_trades if not t["is_winner"] and t["exit_reason"] == "SL_HIT"]
    if not losers: return []

    loss_patterns = defaultdict(lambda: {"count": 0, "total_loss": 0, "examples": []})

    for t in losers:
        reasons = []

        # 1. Against trend?
        if t["direction"] == "CE" and t["trend"] in ("BEAR", "STRONG_BEAR"):
            reasons.append("COUNTER_TREND_CE_IN_BEAR")
        elif t["direction"] == "PE" and t["trend"] in ("BULL", "STRONG_BULL"):
            reasons.append("COUNTER_TREND_PE_IN_BULL")

        # 2. Low score?
        if t["score"] <= MIN_SCORE:
            reasons.append(f"MARGINAL_SCORE_{t['score']}")

        # 3. Overextended entry? (RSI)
        if t["direction"] == "CE" and t["rsi"] > 60:
            reasons.append(f"OVERBOUGHT_ENTRY_RSI_{t['rsi']:.0f}")
        elif t["direction"] == "PE" and t["rsi"] < 40:
            reasons.append(f"OVERSOLD_ENTRY_RSI_{t['rsi']:.0f}")

        # 4. Low volume?
        if t["vol_ratio"] < 0.8:
            reasons.append(f"LOW_VOLUME_{t['vol_ratio']:.1f}x")

        # 5. Against SuperTrend?
        if t["direction"] == "CE" and not t["st_bull"]:
            reasons.append("AGAINST_SUPERTREND")
        elif t["direction"] == "PE" and t["st_bull"]:
            reasons.append("AGAINST_SUPERTREND")

        # 6. Time of day?
        h = int(t["entry_time"][:2]) if t["entry_time"] else 0
        if h < 10:
            reasons.append("EARLY_SESSION_ENTRY")
        elif h >= 15:
            reasons.append("LATE_SESSION_ENTRY")

        # 7. Setup type that's losing?
        reasons.append(f"SETUP_{t['setup_type']}")

        if not reasons:
            reasons.append("NO_CLEAR_PATTERN")

        t["loss_reasons"] = reasons

        for r in reasons:
            loss_patterns[r]["count"] += 1
            loss_patterns[r]["total_loss"] += abs(t["pnl_rs"])
            if len(loss_patterns[r]["examples"]) < 3:
                loss_patterns[r]["examples"].append(t["id"])

    # Sort by frequency
    sorted_patterns = sorted(loss_patterns.items(), key=lambda x: -x[1]["count"])
    return [{"pattern": k, **v, "avg_loss": v["total_loss"]/max(v["count"],1)}
            for k, v in sorted_patterns]


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT (single file, multiple period sheets)
# ═══════════════════════════════════════════════════════════════════════════════

def export_excel(all_period_results: list, loss_analysis: list, out_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference

    thin=Side(style="thin",color="BFBFBF")
    brd=Border(left=thin,right=thin,top=thin,bottom=thin)
    HF=Font(bold=True,color="1F497D",size=10)
    HBG=PatternFill("solid",fgColor="DCE6F1")
    WF=PatternFill("solid",fgColor="E2EFDA")
    LF=PatternFill("solid",fgColor="FCE4D6")
    GF=Font(bold=True,color="375623",size=10)
    RF=Font(bold=True,color="9C0006",size=10)

    wb=Workbook()

    # ── MASTER DASHBOARD ──────────────────────────────────────────
    ws=wb.active; ws.title="Dashboard"
    ws["A1"]="SCALPER PRO v3 — Institutional Backtest Report"; ws["A1"].font=Font(bold=True,size=14,color="1F497D")
    ws.merge_cells("A1:J1")

    r=3
    headers=["Period","Indices","Trades","Win","Loss","Win%","Net P&L","PF","Best Setup","Data Source"]
    for ci,h in enumerate(headers,1):
        c=ws.cell(r,ci,h); c.font=HF; c.fill=HBG; c.border=brd

    for label,sd,ed,results,data_src in all_period_results:
        r+=1
        at=[t for _,trades,_,_,_ in results for t in trades]
        w=sum(1 for t in at if t["is_winner"]); net=sum(t["pnl_rs"] for t in at)
        gw=sum(t["pnl_rs"] for t in at if t["is_winner"])
        gl=abs(sum(t["pnl_rs"] for t in at if not t["is_winner"]))
        wr=w/len(at)*100 if at else 0; pf=gw/gl if gl>0 else 0
        setups=defaultdict(float)
        for t in at: setups[t["setup_type"]]+=t["pnl_rs"]
        best_setup=max(setups.items(),key=lambda x:x[1])[0] if setups else ""
        idxs=", ".join(set(t["index"] for t in at))
        vals=[label,idxs,len(at),w,len(at)-w,f"{wr:.0f}%",f"Rs {net:+,.0f}",f"{pf:.2f}",best_setup,data_src]
        for ci,v in enumerate(vals,1):
            c=ws.cell(r,ci,v); c.border=brd
            if ci==7: c.font=GF if net>=0 else RF

    for col,w in zip("ABCDEFGHIJ",[22,18,8,6,6,8,16,8,16,12]):
        ws.column_dimensions[col].width=w

    # ── PER-PERIOD TRADE SHEETS ───────────────────────────────────
    tcols=["ID","Date","Time","Index","Dir","Setup","Entry","Target","SL","Exit","ExitTime",
           "Reason","Hold","PnL pts","PnL Rs","Score","Trend","RSI","BB%b","ST","VolR","Win","Reasons"]

    for label,sd,ed,results,_ in all_period_results:
        at=[t for _,trades,_,_,_ in results for t in trades]
        if not at: continue
        sname=label[:28].replace("/","-")  # Excel sheet name limit
        wt=wb.create_sheet(sname)
        for ci,h in enumerate(tcols,1):
            c=wt.cell(1,ci,h); c.font=HF; c.fill=HBG; c.border=brd
        wt.auto_filter.ref=f"A1:W{len(at)+1}"; wt.freeze_panes="A2"

        for ri,t in enumerate(at,2):
            fill=WF if t["is_winner"] else LF
            vals=[t["id"],t["entry_date"],t["entry_time"],t["index"],t["direction"],t["setup_type"],
                  round(t["entry_spot"],0),round(t["target_spot"],0),round(t["sl_spot"],0),
                  round(t["exit_spot"],0) if t["exit_spot"] else "",t["exit_time"] or "",
                  t["exit_reason"] or "",t["holding_bars"],round(t["pnl_pts"],1),round(t["pnl_rs"],0),
                  f"{t['score']}/{t.get('max_score',MAX_SCORE)}",t["trend"],
                  round(t["rsi"],1),round(t["bb_pct_b"],2),"G" if t["st_bull"] else "R",
                  round(t["vol_ratio"],2),"WIN" if t["is_winner"] else "LOSS",
                  (t["reasons"][:100] if isinstance(t["reasons"],str) else str(t["reasons"])[:100])]
            for ci,v in enumerate(vals,1):
                c=wt.cell(ri,ci,v); c.fill=fill; c.border=brd

        for ci,w in enumerate([10,10,6,8,4,14,9,9,9,9,6,10,5,8,10,7,12,5,5,3,6,5,60],1):
            wt.column_dimensions[wt.cell(1,ci).column_letter].width=w

    # ── LOSS ANALYSIS (self-learning) ─────────────────────────────
    wl=wb.create_sheet("Loss Analysis")
    wl["A1"]="SELF-LEARNING: Why Trades Failed"; wl["A1"].font=Font(bold=True,size=13,color="9C0006")
    wl.merge_cells("A1:E1")

    r=3
    for ci,h in enumerate(["Pattern","Count","Total Loss","Avg Loss","Examples"],1):
        c=wl.cell(r,ci,h); c.font=HF; c.fill=HBG; c.border=brd

    for lp in loss_analysis[:20]:
        r+=1
        vals=[lp["pattern"],lp["count"],f"Rs {lp['total_loss']:,.0f}",
              f"Rs {lp['avg_loss']:,.0f}",", ".join(lp["examples"])]
        for ci,v in enumerate(vals,1):
            c=wl.cell(r,ci,v); c.border=brd

    # Recommendations
    r+=2
    wl.cell(r,1,"RECOMMENDATIONS").font=Font(bold=True,size=12,color="1F497D")
    r+=1
    recs=[]
    for lp in loss_analysis[:5]:
        p=lp["pattern"]
        if "COUNTER_TREND" in p: recs.append(f"AVOID {p}: Never trade CE in BEAR or PE in BULL trend")
        elif "AGAINST_SUPERTREND" in p: recs.append(f"AVOID {p}: SuperTrend must align with direction")
        elif "MARGINAL_SCORE" in p: recs.append(f"RAISE MIN_SCORE: Trades at score {MIN_SCORE} have high failure rate")
        elif "LOW_VOLUME" in p: recs.append(f"SKIP LOW VOLUME: Require vol_ratio >= 1.0 minimum")
        elif "OVERBOUGHT" in p or "OVERSOLD" in p: recs.append(f"FIX RSI FILTER: {p}")
        elif "EARLY_SESSION" in p: recs.append(f"MOVE NO_TRADE_BEFORE to 10:00 (9:45 still has opening noise)")
        elif "LATE_SESSION" in p: recs.append(f"MOVE NO_TRADE_AFTER to 14:45 (15:00 has EOD decay)")
        else: recs.append(f"Investigate: {p} ({lp['count']} losses, avg Rs {lp['avg_loss']:,.0f})")
    for rec in recs:
        wl.cell(r,1,f"  → {rec}").font=Font(size=10)
        r+=1

    for col,w in zip("ABCDE",[32,8,14,14,30]):
        wl.column_dimensions[col].width=w

    # ── EQUITY CURVE ──────────────────────────────────────────────
    all_t=[t for _,_,_,results,_ in all_period_results for _,trades,_,_,_ in results for t in trades]
    if all_t:
        we=wb.create_sheet("Equity Curve")
        we.cell(1,1,"Trade#").font=HF; we.cell(1,2,"Equity").font=HF; we.cell(1,3,"Drawdown").font=HF
        cum=0; peak=0
        for i,t in enumerate(all_t):
            cum+=t["pnl_rs"]; peak=max(peak,cum)
            we.cell(i+2,1,i+1); we.cell(i+2,2,round(cum,0)); we.cell(i+2,3,round(cum-peak,0))
        if len(all_t)>3:
            ch=LineChart(); ch.title="Equity Curve"; ch.style=10; ch.width=28; ch.height=14
            ch.add_data(Reference(we,min_col=2,min_row=1,max_row=len(all_t)+1),titles_from_data=True)
            ch.set_categories(Reference(we,min_col=1,min_row=2,max_row=len(all_t)+1))
            we.add_chart(ch,"E2")

    wb.save(out_path)
    print(f"\n  Excel saved: {out_path}")


# ═══════════════════════════════════════════════════════════════════════════════
# FIND TRADING DAYS
# ═══════════════════════════════════════════════════════════════════════════════

def find_last_trading_day(fifteen_df, target=None):
    """Find most recent trading day in the data. If target is given, find it or previous."""
    if fifteen_df is None or len(fifteen_df)==0: return None
    dates=sorted(fifteen_df["datetime"].dt.date.unique())
    if target:
        valid=[d for d in dates if d<=target]
        return valid[-1] if valid else None
    return dates[-1] if dates else None


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    indices = list(INDEX_PARAMS.keys())
    today = date.today()

    print("\n" + "=" * 72)
    print("  SCALPER PRO v3 — Complete Institutional Backtest")
    print(f"  Indices : {', '.join(indices)}")
    print(f"  Scoring : {MAX_SCORE}-point ({MIN_SCORE} minimum)")
    print(f"  Setups  : BOX | EMA_PB | ST_FLIP | VWAP | SR_BOUNCE | LIQ_SWEEP")
    print("=" * 72)

    # ── Fetch data for all indices ────────────────────────────────
    data_cache = {}
    data_source = "Dhan+Yahoo"
    for index in indices:
        print(f"\n  [{index}] Fetching data...")
        daily, fifteen = fetch_data(index, days_back=180)
        data_cache[index] = (daily, fifteen)
        d_len = len(daily) if daily is not None else 0
        f_len = len(fifteen) if fifteen is not None else 0
        if d_len == 0 and f_len == 0:
            data_source = "INSUFFICIENT"
            print(f"    ⚠️ NO DATA for {index} — check Dhan credentials in .env")
        else:
            print(f"    Daily={d_len}  15min={f_len}")

    # ── Define periods ────────────────────────────────────────────
    # Find the actual last trading day from the data
    sample_fifteen = next((v[1] for v in data_cache.values() if v[1] is not None and len(v[1])>0), None)
    last_td = find_last_trading_day(sample_fifteen, today) or today

    periods = [
        (f"Today ({last_td})", last_td, last_td),
        (f"Yesterday ({last_td - timedelta(days=1)})", last_td - timedelta(days=1), last_td - timedelta(days=1)),
        (f"This Week ({last_td - timedelta(days=last_td.weekday())} to {last_td})",
         last_td - timedelta(days=last_td.weekday()), last_td),
        (f"This Month ({last_td.replace(day=1)} to {last_td})",
         last_td.replace(day=1), last_td),
        (f"6 Months ({last_td - timedelta(days=180)} to {last_td})",
         last_td - timedelta(days=180), last_td),
    ]

    all_period_results = []
    all_trades_flat = []

    for period_idx, (label, sd, ed) in enumerate(periods):
        print(f"\n{'─'*72}")
        print(f"  PERIOD {period_idx+1}/5: {label}")
        print(f"{'─'*72}")

        period_results = []
        for index in indices:
            daily, fifteen = data_cache[index]
            params = INDEX_PARAMS[index]
            if daily is None or fifteen is None or len(fifteen) < 50:
                period_results.append((index, [], 0, 0, params))
                continue
            trades, signals, bars = run_period(index, params, daily, fifteen, sd, ed)
            period_results.append((index, trades, signals, bars, params))

            w=sum(1 for t in trades if t["is_winner"]); net=sum(t["pnl_rs"] for t in trades)
            wr=w/len(trades)*100 if trades else 0
            all_trades_flat.extend(trades)
            print(f"  [{index:12s}] Trades={len(trades):3d}  Win={w}({wr:4.0f}%)  P&L=Rs{net:>+10,.0f}  Signals={signals}")

        at=[t for _,trades,_,_,_ in period_results for t in trades]
        tw=sum(1 for t in at if t["is_winner"]); tn=sum(t["pnl_rs"] for t in at)
        twr=tw/len(at)*100 if at else 0
        print(f"\n  TOTAL: {len(at)} trades | Win: {tw} ({twr:.0f}%) | P&L: Rs {tn:+,.0f}")

        all_period_results.append((label, sd, ed, period_results, data_source))

        # Prompt before next period
        if period_idx < len(periods) - 1:
            try:
                ans = input(f"\n  Press ENTER to run next period, or 'q' to stop: ").strip().lower()
                if ans == 'q':
                    print("  Stopping early.")
                    break
            except (EOFError, KeyboardInterrupt):
                break

    # ── Loss analysis ─────────────────────────────────────────────
    print(f"\n{'─'*72}")
    print("  SELF-LEARNING LOSS ANALYSIS")
    print(f"{'─'*72}")
    loss_patterns = analyze_losses(all_trades_flat)
    for lp in loss_patterns[:8]:
        print(f"  {lp['pattern']:35s}  {lp['count']:3d} trades  Avg loss Rs {lp['avg_loss']:>8,.0f}")

    # ── Export Excel ──────────────────────────────────────────────
    out_path = os.path.join(_this_dir, f"InstitutionalBT_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
    export_excel(all_period_results, loss_patterns, out_path)
    print(f"\n{'='*72}")
    print(f"  DONE. Excel: {out_path}")
    print(f"{'='*72}\n")


if __name__ == "__main__":
    main()
