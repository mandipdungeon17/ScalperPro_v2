"""NIFTY backtest: Yesterday (1-Apr-2026) + Last 1 Month with full indicator stack."""
import sys, os, warnings, logging
warnings.filterwarnings("ignore")
sys.path.insert(0, os.path.dirname(__file__))

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s  %(levelname)-7s  %(message)s",
                    datefmt="%H:%M:%S")
logger = logging.getLogger(__name__)

from datetime import date, timedelta, datetime
from collections import Counter

TODAY    = date(2026, 4, 2)
YEST     = date(2026, 4, 1)
M1_START = TODAY - timedelta(days=30)


def main():
    from scalper.data.fetcher import DataFetcher
    from scalper.backtest.nifty_backtest import NiftyBacktest, analyze_long_term_levels

    print("=" * 68)
    print("  SCALPER PRO v2 -- NIFTY BACKTEST  [Yesterday + 1 Month]")
    print("  Indicators: EMA20, 5EMA, 13EMA H/L, SuperTrend(10,3),")
    print("              RSI(14), Volume, Bollinger Bands(20,2)")
    print("  Signal gate: need 3/7 indicators aligned to enter")
    print("=" * 68)

    fetcher = DataFetcher()
    bt = NiftyBacktest(fetcher)

    print("[1/4] Fetching data (daily 2yr + 15-min 90d)...")
    bt.fetch_all_data()

    logging.getLogger("scalper.core.index_levels").setLevel(logging.WARNING)
    logging.getLogger("scalper.core.premarket_analysis").setLevel(logging.WARNING)

    print("[2/4] Running backtests...")
    r_yest = bt._run_intraday(
        label="1-Apr-2026 (Yesterday)",
        start_date=YEST, end_date=YEST,
    )
    r_1m = bt._run_intraday(
        label="1-Month (Mar-3 to Apr-1)",
        start_date=M1_START, end_date=YEST,
    )
    r_yest.failure_analysis = bt._analyze_failures(r_yest.trades)
    r_1m.failure_analysis   = bt._analyze_failures(r_1m.trades)
    results = [r_yest, r_1m]

    print("[3/4] Analysing 2-year NIFTY key levels...")
    levels_df, levels_summary = analyze_long_term_levels(fetcher)

    out_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        f"NIFTY_Yest_1M_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
    )
    print(f"[4/4] Exporting to Excel: {out_path}")
    export_excel(results, levels_df, levels_summary, out_path)

    print_summary(results, out_path)
    return results, out_path


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL
# ─────────────────────────────────────────────────────────────────────────────

TRADE_COLS = [
    "ID", "Date", "Time", "Dir", "Level", "Level TF", "Touches", "Strength",
    "Zone", "Dist ATR", "Entry", "Target", "SL", "Exit", "Exit Time",
    "Exit Reason", "Hold", "Idx Pts", "Opt Pts", "P&L (Rs)", "W/L",
    "Score", "Score Detail", "RSI@Entry", "ST Bull?", "BB pct_b", "Quality",
]


def export_excel(results, levels_df, levels_summary, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference

    thin = Side(style="thin", color="BFBFBF")
    BRD  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hf():
        return Font(bold=True, color="1F497D", size=10)

    def hb():
        return PatternFill("solid", fgColor="DCE6F1")

    WIN  = PatternFill("solid", fgColor="E2EFDA")
    LOSS = PatternFill("solid", fgColor="FCE4D6")
    SUP  = PatternFill("solid", fgColor="DAEEF3")
    RES  = PatternFill("solid", fgColor="F2DCDB")

    def hrow(ws, cols, row_num=1):
        for ci, c in enumerate(cols, 1):
            x = ws.cell(row_num, ci, c)
            x.font = hf(); x.fill = hb(); x.border = BRD
            x.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.row_dimensions[row_num].height = 28

    wb = Workbook()

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "NIFTY Backtest -- Yesterday + 1-Month  (Multi-Indicator System)"
    ws["A1"].font = Font(bold=True, size=13, color="1F497D")
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 26
    ws["A2"] = (
        "Filters: S/R levels (2+ touches, strength>=0.30, max 4 touches), EMA20 trend, "
        "rejection candle, THEN 7-indicator gate: 5EMA + 13EMA + 13EMA-H/L + "
        "SuperTrend(10,3) + RSI(14) + Volume + BB(20,2) -- need 3/7 to trade"
    )
    ws["A2"].font = Font(size=9, italic=True, color="595959")
    ws.merge_cells("A2:H2")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 36

    hrow(ws, ["Metric", "Yesterday 1-Apr", "1-Month (Mar3-Apr1)"], row_num=4)
    r_y, r_m = results[0], results[1]
    summary_rows = [
        ("Period",         r_y.start_date,       f"{r_m.start_date} to {r_m.end_date}"),
        ("Timeframe",      r_y.timeframe,         r_m.timeframe),
        ("Total Bars",     r_y.total_bars,        r_m.total_bars),
        ("Trades Taken",   r_y.total_trades,      r_m.total_trades),
        ("Winners",        r_y.winners,           r_m.winners),
        ("Losers",         r_y.losers,            r_m.losers),
        ("Win Rate",       f"{r_y.win_rate:.1f}%",  f"{r_m.win_rate:.1f}%"),
        ("Net P&L",        f"Rs {r_y.total_pnl:+,.0f}", f"Rs {r_m.total_pnl:+,.0f}"),
        ("Profit Factor",  f"{r_y.profit_factor:.2f}", f"{r_m.profit_factor:.2f}"),
        ("Sharpe Ratio",   f"{r_y.sharpe_ratio:.2f}",  f"{r_m.sharpe_ratio:.2f}"),
        ("Max Drawdown",   f"Rs {r_y.max_drawdown:,.0f}", f"Rs {r_m.max_drawdown:,.0f}"),
        ("Best Trade",     f"Rs {r_y.best_trade_pnl:+,.0f}", f"Rs {r_m.best_trade_pnl:+,.0f}"),
        ("Worst Trade",    f"Rs {r_y.worst_trade_pnl:+,.0f}", f"Rs {r_m.worst_trade_pnl:+,.0f}"),
        ("Avg Hold Bars",  f"{r_y.avg_hold_bars:.1f}",  f"{r_m.avg_hold_bars:.1f}"),
        ("Max Win Streak", r_y.max_win_streak,    r_m.max_win_streak),
        ("Max Loss Streak",r_y.max_loss_streak,   r_m.max_loss_streak),
    ]
    for ri, (k, v1, v2) in enumerate(summary_rows, 5):
        c_k = ws.cell(ri, 1, k)
        c_k.font = Font(bold=True, size=10); c_k.fill = hb(); c_k.border = BRD
        c1 = ws.cell(ri, 2, v1); c1.border = BRD
        c1.alignment = Alignment(horizontal="right")
        c2 = ws.cell(ri, 3, v2); c2.border = BRD
        c2.alignment = Alignment(horizontal="right")
        if k == "Net P&L":
            c1.font = Font(bold=True, color=("375623" if r_y.total_pnl >= 0 else "9C0006"), size=11)
            c2.font = Font(bold=True, color=("375623" if r_m.total_pnl >= 0 else "9C0006"), size=11)

    for col, w in [("A", 24), ("B", 22), ("C", 26), ("D", 14), ("E", 14), ("F", 14)]:
        ws.column_dimensions[col].width = w

    # ── Trade sheets ──────────────────────────────────────────────────────────
    def write_trade_sheet(result, name):
        ws_t = wb.create_sheet(name)
        hrow(ws_t, TRADE_COLS)
        for r_idx, t in enumerate(result.trades, 2):
            fill = WIN if t.is_winner else LOSS
            vals = [
                t.trade_id, t.entry_date, t.entry_time, t.direction,
                t.level_price, t.level_tf, t.level_touches,
                round(t.level_strength, 3), t.proximity_zone,
                round(t.distance_atr, 3), t.entry_spot, t.target_spot, t.sl_spot,
                t.exit_spot, t.exit_time, t.exit_reason, t.holding_bars,
                round(t.index_pnl_pts, 1), round(t.option_pnl_pts, 1),
                round(t.pnl_rupees, 0), "WIN" if t.is_winner else "LOSS",
                t.conf_score, t.conf_breakdown,
                round(t.rsi_at_entry, 1), t.supertrend_bull,
                round(t.bb_pct_b, 2), t.setup_quality,
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws_t.cell(r_idx, ci, v)
                cell.fill = fill; cell.border = BRD
                cell.alignment = Alignment(horizontal="center")
        widths = [7,11,7,5,8,8,7,9,10,8,8,8,8,8,8,12,6,8,8,10,5,6,40,8,7,8,8]
        for ci, w in enumerate(widths, 1):
            ws_t.column_dimensions[ws_t.cell(1, ci).column_letter].width = w
        if result.equity_curve and len(result.equity_curve) > 1:
            ecol = len(TRADE_COLS) + 2
            ws_t.cell(1, ecol, "Equity (Rs)")
            for ri2, v in enumerate(result.equity_curve, 2):
                ws_t.cell(ri2, ecol, round(v, 0))
            chart = LineChart()
            chart.title = f"Equity Curve -- {name}"
            chart.style = 10; chart.width = 24; chart.height = 14
            dr = Reference(ws_t, min_col=ecol, min_row=1,
                           max_row=len(result.equity_curve) + 1)
            chart.add_data(dr, titles_from_data=True)
            ws_t.add_chart(chart, f"A{len(result.trades) + 4}")

    write_trade_sheet(r_y, "Yesterday_1Apr2026")
    write_trade_sheet(r_m, "1Month_Mar3_Apr1")

    # ── Failure + indicator analysis ──────────────────────────────────────────
    ws_f = wb.create_sheet("Analysis")
    ws_f["A1"] = "Failure & Indicator Score Analysis"
    ws_f["A1"].font = Font(bold=True, size=13, color="1F497D")
    cur_row = 3
    for result in results:
        ws_f.cell(cur_row, 1, f"Period: {result.label}").font = Font(bold=True, size=11, color="1F497D")
        cur_row += 1
        if result.failure_analysis:
            hrow(ws_f, ["Failure Reason", "Count"], row_num=cur_row)
            cur_row += 1
            for reason, cnt in sorted(result.failure_analysis.items(), key=lambda x: -x[1]):
                ws_f.cell(cur_row, 1, reason).border = BRD
                ws_f.cell(cur_row, 2, cnt).border = BRD
                cur_row += 1
        else:
            ws_f.cell(cur_row, 1, "No trades to analyze").font = Font(italic=True, color="595959")
            cur_row += 1

        if result.trades:
            cur_row += 1
            ws_f.cell(cur_row, 1, "Indicator Score vs Win Rate").font = Font(bold=True, size=10)
            cur_row += 1
            hrow(ws_f, ["Score (out of 7)", "Trades", "Wins", "Losses", "Win%", "Avg P&L"], row_num=cur_row)
            cur_row += 1
            scores = Counter(t.conf_score for t in result.trades)
            for sc in sorted(scores):
                sc_trades = [t for t in result.trades if t.conf_score == sc]
                wins = sum(1 for t in sc_trades if t.is_winner)
                losses = len(sc_trades) - wins
                wr = 100 * wins / len(sc_trades) if sc_trades else 0
                avg_pnl = sum(t.pnl_rupees for t in sc_trades) / len(sc_trades)
                ws_f.cell(cur_row, 1, sc).border = BRD
                ws_f.cell(cur_row, 2, len(sc_trades)).border = BRD
                ws_f.cell(cur_row, 3, wins).border = BRD
                ws_f.cell(cur_row, 4, losses).border = BRD
                ws_f.cell(cur_row, 5, f"{wr:.0f}%").border = BRD
                ws_f.cell(cur_row, 6, f"Rs {avg_pnl:+,.0f}").border = BRD
                cur_row += 1
        cur_row += 2

    ws_f.column_dimensions["A"].width = 48
    for col in "BCDEF":
        ws_f.column_dimensions[col].width = 12

    # ── Key Levels sheet ──────────────────────────────────────────────────────
    ws_l = wb.create_sheet("NIFTY_Key_Levels")
    if levels_df is not None and len(levels_df) > 0:
        hrow(ws_l, list(levels_df.columns))
        for r_idx, row_data in enumerate(levels_df.itertuples(index=False), 2):
            ltype = str(getattr(row_data, "Type", "")).upper()
            fill = SUP if ltype == "SUPPORT" else RES
            for ci, v in enumerate(row_data, 1):
                cell = ws_l.cell(r_idx, ci, v)
                cell.fill = fill; cell.border = BRD
                cell.alignment = Alignment(horizontal="center")
        for ci in range(1, len(levels_df.columns) + 1):
            ws_l.column_dimensions[ws_l.cell(1, ci).column_letter].width = 16

    wb.save(out_path)
    logger.info(f"Excel saved: {out_path}")


# ─────────────────────────────────────────────────────────────────────────────
# CONSOLE SUMMARY
# ─────────────────────────────────────────────────────────────────────────────

def print_summary(results, out_path):
    print()
    print("=" * 68)
    print(f"  {'PERIOD':<30} {'TRD':>4} {'WIN%':>6} {'P&L':>11} {'PF':>5} {'SHRP':>6}")
    print("-" * 68)
    for r in results:
        sign = "+" if r.total_pnl >= 0 else ""
        print(f"  {r.label:<30} {r.total_trades:>4} "
              f"{r.win_rate:>5.1f}%  Rs{sign}{r.total_pnl:>8,.0f} "
              f"{r.profit_factor:>5.2f} {r.sharpe_ratio:>6.2f}")
    print("=" * 68)

    for r in results:
        if not r.trades:
            print(f"\n  {r.label}: 0 trades -- no qualifying setup")
            print("    (NIFTY not near any validated S/R level on this day/period)")
            continue
        print(f"\n  {r.label} -- {r.total_trades} trades:")
        print(f"  {'#':<7} {'Date':<12} {'T':<6} {'D':<3} {'Lvl':>7} "
              f"{'Sc':>4} {'RSI':>5} {'BB':>5} {'Exit':<12} {'P&L':>10}")
        print("  " + "-" * 72)
        for t in r.trades:
            pstr = f"Rs {t.pnl_rupees:+,.0f}"
            sc_str = f"{t.conf_score}/7"
            print(f"  {t.trade_id:<7} {t.entry_date:<12} {t.entry_time:<6} "
                  f"{t.direction:<3} {t.level_price:>7.0f} "
                  f"{sc_str:>4} {t.rsi_at_entry:>5.1f} {t.bb_pct_b:>5.2f} "
                  f"{t.exit_reason:<12} {pstr:>10}")

    print(f"\n  Excel: {out_path}\n")


if __name__ == "__main__":
    main()
