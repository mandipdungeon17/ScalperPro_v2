"""
=============================================================================
SCALPER PRO v2 — NIFTY Backtest Runner
=============================================================================
Runs all 4 backtest periods and exports results to Excel.

Usage:
    python run_nifty_backtest.py

Output:
    D:/Trading/ScalperPro_v2/NIFTY_Backtest_<date>.xlsx
=============================================================================
"""

import sys
import os
import logging
import warnings
import numpy as np
import pandas as pd
from datetime import datetime
from typing import List

warnings.filterwarnings("ignore")
sys.path.insert(0, os.path.dirname(__file__))

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)


def run_backtest():
    from scalper.data.fetcher import DataFetcher
    from scalper.backtest.nifty_backtest import NiftyBacktest, analyze_long_term_levels, PeriodResult

    print("\n" + "=" * 70)
    print("  SCALPER PRO v2 — NIFTY BACKTEST")
    print("=" * 70)

    fetcher = DataFetcher()
    bt = NiftyBacktest(fetcher)

    print("[1/3] Running backtests (Today / 1-Month / 6-Month / 1-Year)...")
    results = bt.run_all()

    print("[2/3] Analysing 2-year NIFTY levels...")
    levels_df, levels_summary = analyze_long_term_levels(fetcher)

    print("[3/3] Exporting to Excel...")
    out_path = os.path.join(
        os.path.dirname(__file__),
        f"NIFTY_Backtest_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )
    export_excel(results, levels_df, levels_summary, out_path)

    # Print console summary
    print_console_summary(results)
    print(f"\nDone. Excel saved: {out_path}\n")
    return results, out_path


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def export_excel(
    results: List,
    levels_df: pd.DataFrame,
    levels_summary: pd.DataFrame,
    out_path: str,
):
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.series import SeriesLabel

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Colour palette ─────────────────────────────────────────────────────
    C_HEADER   = "1F3864"   # dark navy
    C_SUBHDR   = "2E75B6"   # medium blue
    C_WIN      = "E2EFDA"   # light green
    C_LOSS     = "FCE4D6"   # light red
    C_NEUTRAL  = "FFF2CC"   # light yellow
    C_LEVEL_S  = "DAEEF3"   # support — light blue
    C_LEVEL_R  = "F2DCDB"   # resistance — light red
    C_ROW_ALT  = "F5F5F5"   # alternating row

    def hdr_font(size=11, bold=True, color="FFFFFF"):
        return Font(name="Calibri", size=size, bold=bold, color=color)

    def body_font(size=10, bold=False, color="000000"):
        return Font(name="Calibri", size=size, bold=bold, color=color)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def center(wrap=False):
        return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

    def left(wrap=False):
        return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

    thin = Side(style="thin", color="CCCCCC")
    def border():
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def pct_fmt(val):
        return f"{val:.1f}%"

    def inr_fmt(val):
        sign = "+" if val >= 0 else ""
        return f"₹{sign}{val:,.0f}"

    def set_col_widths(ws, widths):
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    def write_header_row(ws, row_num, headers, col_fill=C_HEADER, height=20):
        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=row_num, column=ci, value=h)
            cell.font      = hdr_font()
            cell.fill      = fill(col_fill)
            cell.alignment = center()
            cell.border    = border()
        ws.row_dimensions[row_num].height = height

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 1: SUMMARY (all 4 periods)
    # ══════════════════════════════════════════════════════════════════════
    ws_sum = wb.create_sheet("Summary")
    ws_sum.sheet_view.showGridLines = False

    # Title
    ws_sum.merge_cells("A1:L1")
    t = ws_sum["A1"]
    t.value     = "NIFTY BACKTEST — ALL PERIODS SUMMARY"
    t.font      = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t.fill      = fill(C_HEADER)
    t.alignment = center()
    ws_sum.row_dimensions[1].height = 28

    ws_sum.merge_cells("A2:L2")
    t2 = ws_sum["A2"]
    t2.value     = f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}  |  Algorithm: L1 S/R Proximity + L3 Greek Strike  |  NIFTY Lot Size: 75"
    t2.font      = body_font(9, color="555555")
    t2.alignment = center()

    # Headers
    headers = [
        "Period", "Timeframe", "Start Date", "End Date",
        "Total Trades", "Winners", "Losers", "Win Rate",
        "Total P&L (₹)", "Profit Factor", "Sharpe Ratio", "Max Drawdown (₹)",
        "Avg Win (₹)", "Avg Loss (₹)", "Best Trade (₹)", "Worst Trade (₹)",
        "Max Win Streak", "Max Loss Streak", "Avg Hold Bars",
    ]
    write_header_row(ws_sum, 4, headers)

    for ri, r in enumerate(results, start=5):
        row_fill = fill(C_ROW_ALT) if ri % 2 == 0 else fill("FFFFFF")
        vals = [
            r.label, r.timeframe, r.start_date, r.end_date,
            r.total_trades, r.winners, r.losers, f"{r.win_rate:.1f}%",
            r.total_pnl, r.profit_factor, r.sharpe_ratio, r.max_drawdown,
            r.avg_win, r.avg_loss, r.best_trade_pnl, r.worst_trade_pnl,
            r.max_win_streak, r.max_loss_streak, r.avg_hold_bars,
        ]
        for ci, v in enumerate(vals, start=1):
            cell = ws_sum.cell(row=ri, column=ci, value=v)
            cell.font      = body_font()
            cell.alignment = center()
            cell.border    = border()

            # Colour P&L column
            if ci == 9:  # Total P&L
                if isinstance(v, (int, float)):
                    cell.fill = fill(C_WIN if v >= 0 else C_LOSS)
                    cell.font = body_font(bold=True, color="1A6A1A" if v >= 0 else "B00000")

    set_col_widths(ws_sum, [12, 10, 12, 12, 12, 9, 9, 9, 14, 13, 12, 16, 13, 13, 14, 15, 14, 15, 14])

    # Monthly P&L mini-table below summary
    all_months = set()
    for r in results:
        all_months.update(r.monthly_pnl.keys())
    all_months = sorted(all_months)

    if all_months:
        row_start = 5 + len(results) + 2
        ws_sum.cell(row=row_start - 1, column=1).value = "MONTHLY P&L BREAKDOWN (₹)"
        ws_sum.cell(row=row_start - 1, column=1).font  = hdr_font(11, color="000000")
        ws_sum.cell(row=row_start - 1, column=1).fill  = fill(C_SUBHDR)

        month_hdrs = ["Month"] + [r.label for r in results]
        write_header_row(ws_sum, row_start, month_hdrs, col_fill=C_SUBHDR)

        for mi, mo in enumerate(all_months, start=row_start + 1):
            ws_sum.cell(row=mi, column=1, value=mo).alignment = center()
            for ci, r in enumerate(results, start=2):
                val = r.monthly_pnl.get(mo, 0)
                cell = ws_sum.cell(row=mi, column=ci, value=val)
                cell.alignment = center()
                cell.fill = fill(C_WIN if val > 0 else C_LOSS if val < 0 else "FFFFFF")
                cell.font = body_font(bold=True, color="1A6A1A" if val > 0 else "B00000" if val < 0 else "000000")

    # ══════════════════════════════════════════════════════════════════════
    # SHEETS 2-5: Individual period trade logs
    # ══════════════════════════════════════════════════════════════════════
    trade_headers = [
        "Trade #", "Entry Date", "Entry Time", "Direction",
        "Level Type", "Level Price", "Level TF", "Strength", "Touches",
        "Proximity Zone", "Dist ATR", "Setup Grade",
        "Entry Spot", "SL", "Target",
        "Exit Date", "Exit Time", "Exit Spot", "Exit Reason",
        "Holding Bars", "Index P&L (pts)", "Option P&L (pts)", "P&L (₹)", "Result",
    ]

    for r in results:
        ws = wb.create_sheet(r.label)
        ws.sheet_view.showGridLines = False

        # Title bar
        ws.merge_cells(f"A1:{get_column_letter(len(trade_headers))}1")
        tc = ws["A1"]
        tc.value     = f"NIFTY  |  {r.label}  |  {r.start_date} → {r.end_date}  |  {r.timeframe}"
        tc.font      = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        tc.fill      = fill(C_HEADER)
        tc.alignment = center()
        ws.row_dimensions[1].height = 24

        # Stats row
        stats_txt = (
            f"Trades: {r.total_trades}  |  "
            f"Win Rate: {r.win_rate:.1f}%  |  "
            f"Net P&L: {inr_fmt(r.total_pnl)}  |  "
            f"PF: {r.profit_factor:.2f}  |  "
            f"Sharpe: {r.sharpe_ratio:.2f}  |  "
            f"Max DD: {inr_fmt(-r.max_drawdown)}  |  "
            f"Avg Win: {inr_fmt(r.avg_win)}  |  "
            f"Avg Loss: {inr_fmt(r.avg_loss)}"
        )
        ws.merge_cells(f"A2:{get_column_letter(len(trade_headers))}2")
        sc = ws["A2"]
        sc.value     = stats_txt
        sc.font      = body_font(9, color="333333")
        sc.fill      = fill("EBF3FB")
        sc.alignment = left()

        write_header_row(ws, 4, trade_headers)

        for ti, t in enumerate(r.trades, start=5):
            alt = fill(C_ROW_ALT) if ti % 2 == 0 else fill("FFFFFF")
            pnl_color = fill(C_WIN) if t.pnl_rupees > 0 else fill(C_LOSS)

            row_vals = [
                t.trade_id, t.entry_date, t.entry_time, t.direction,
                t.level_type, t.level_price, t.level_tf,
                round(t.level_strength, 3), t.level_touches,
                t.proximity_zone, round(t.distance_atr, 2), t.setup_quality,
                round(t.entry_spot, 2), round(t.sl_spot, 2), round(t.target_spot, 2),
                t.exit_date, t.exit_time, round(t.exit_spot, 2), t.exit_reason,
                t.holding_bars,
                round(t.index_pnl_pts, 2), round(t.option_pnl_pts, 2),
                round(t.pnl_rupees, 2),
                "WIN" if t.is_winner else "LOSS",
            ]
            for ci, v in enumerate(row_vals, start=1):
                cell = ws.cell(row=ti, column=ci, value=v)
                cell.font      = body_font()
                cell.border    = border()
                cell.alignment = center()

                # Colour by win/loss for result column and P&L
                if ci == 23:  # P&L ₹
                    cell.fill = pnl_color
                    cell.font = body_font(bold=True,
                                          color="1A6A1A" if t.pnl_rupees > 0 else "B00000")
                elif ci == 24:  # WIN/LOSS
                    cell.fill = pnl_color
                elif ci == 4:  # Direction
                    cell.fill = fill("D9F2D9") if v == "CE" else fill("FCE4D6")
                elif ci == 5:  # Level type
                    cell.fill = fill(C_LEVEL_S) if v == "SUPPORT" else fill(C_LEVEL_R)
                else:
                    cell.fill = alt

        set_col_widths(ws, [
            9, 12, 10, 9, 11, 11, 9, 9, 8,
            14, 9, 11,
            12, 12, 12,
            12, 10, 12, 12,
            11, 14, 15, 12, 8,
        ])

        # Equity curve chart (if enough trades)
        if len(r.trades) >= 3 and r.equity_curve:
            equity_row_start = len(r.trades) + 7
            ws.cell(row=equity_row_start, column=1, value="Bar").font = hdr_font(color="000000")
            ws.cell(row=equity_row_start, column=2, value="Equity (₹)").font = hdr_font(color="000000")
            for ei, val in enumerate(r.equity_curve):
                ws.cell(row=equity_row_start + 1 + ei, column=1, value=ei)
                ws.cell(row=equity_row_start + 1 + ei, column=2, value=val)

            chart = LineChart()
            chart.title  = f"Equity Curve — {r.label}"
            chart.style  = 10
            chart.y_axis.title = "Cumulative P&L (₹)"
            chart.x_axis.title = "Trade #"
            chart.width  = 20
            chart.height = 10

            data_ref = Reference(ws, min_col=2, min_row=equity_row_start,
                                 max_row=equity_row_start + len(r.equity_curve))
            chart.add_data(data_ref, titles_from_data=True)
            chart.series[0].graphicalProperties.line.solidFill = "2E75B6"
            ws.add_chart(chart, f"A{equity_row_start + len(r.equity_curve) + 3}")

    # ══════════════════════════════════════════════════════════════════════
    # SHEET: Failure Analysis
    # ══════════════════════════════════════════════════════════════════════
    ws_fail = wb.create_sheet("Failure Analysis")
    ws_fail.sheet_view.showGridLines = False

    ws_fail.merge_cells("A1:H1")
    ws_fail["A1"].value     = "NIFTY BACKTEST — FAILURE ANALYSIS & SELF-REVIEW"
    ws_fail["A1"].font      = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws_fail["A1"].fill      = fill(C_HEADER)
    ws_fail["A1"].alignment = center()
    ws_fail.row_dimensions[1].height = 26

    row = 3
    for r in results:
        losing_trades = [t for t in r.trades if not t.is_winner]
        if not losing_trades:
            continue

        # Section header
        ws_fail.cell(row=row, column=1, value=f"Period: {r.label}  ({len(losing_trades)} losses / {r.total_trades} total)").font = hdr_font(11, color="000000")
        ws_fail.cell(row=row, column=1).fill = fill(C_SUBHDR)
        ws_fail.row_dimensions[row].height = 18
        row += 1

        # Failure reason breakdown
        write_header_row(ws_fail, row,
                         ["Failure Category", "Count", "% of Losses", "Fix / Insight"],
                         col_fill="4472C4")
        row += 1
        total_losses = len(losing_trades)
        for reason, cnt in r.failure_analysis.items():
            pct = cnt / total_losses * 100
            insight = _failure_insight(reason)
            ws_fail.cell(row=row, column=1, value=reason).alignment = left(wrap=True)
            ws_fail.cell(row=row, column=2, value=cnt)
            ws_fail.cell(row=row, column=3, value=f"{pct:.1f}%")
            ws_fail.cell(row=row, column=4, value=insight).alignment = left(wrap=True)
            ws_fail.row_dimensions[row].height = 30
            for ci in range(1, 5):
                ws_fail.cell(row=row, column=ci).font   = body_font()
                ws_fail.cell(row=row, column=ci).border = border()
                ws_fail.cell(row=row, column=ci).fill   = fill(C_LOSS if pct > 30 else C_NEUTRAL if pct > 15 else "FFFFFF")
            row += 1

        # Exit reason stats
        row += 1
        ws_fail.cell(row=row, column=1, value="Exit Reason Breakdown (all trades)").font = hdr_font(10, color="000000")
        ws_fail.cell(row=row, column=1).fill = fill("BDD7EE")
        row += 1
        write_header_row(ws_fail, row,
                         ["Exit Reason", "Count", "Avg P&L (₹)", "Win Rate"],
                         col_fill="5B9BD5")
        row += 1
        exit_groups = {}
        for t in r.trades:
            eg = exit_groups.setdefault(t.exit_reason, [])
            eg.append(t.pnl_rupees)
        for reason, pnls in sorted(exit_groups.items()):
            wins = sum(1 for p in pnls if p > 0)
            ws_fail.cell(row=row, column=1, value=reason)
            ws_fail.cell(row=row, column=2, value=len(pnls))
            ws_fail.cell(row=row, column=3, value=round(np.mean(pnls), 2) if pnls else 0)
            ws_fail.cell(row=row, column=4, value=f"{wins/len(pnls)*100:.1f}%" if pnls else "0%")
            for ci in range(1, 5):
                ws_fail.cell(row=row, column=ci).font   = body_font()
                ws_fail.cell(row=row, column=ci).border = border()
                ws_fail.cell(row=row, column=ci).alignment = center()
            row += 1

        row += 2  # gap between periods

    # Quality grade analysis
    row += 1
    ws_fail.cell(row=row, column=1, value="SETUP QUALITY GRADE ANALYSIS (all periods combined)").font = hdr_font(11, color="000000")
    ws_fail.cell(row=row, column=1).fill = fill(C_SUBHDR)
    row += 1
    write_header_row(ws_fail, row,
                     ["Grade", "Trades", "Winners", "Win Rate", "Avg P&L (₹)", "Total P&L (₹)"],
                     col_fill="4472C4")
    row += 1
    all_trades = [t for r in results for t in r.trades]
    for grade in ["A+", "A", "B", "C"]:
        grp = [t for t in all_trades if t.setup_quality == grade]
        if not grp:
            continue
        wins   = sum(1 for t in grp if t.is_winner)
        pnls   = [t.pnl_rupees for t in grp]
        avg_p  = round(np.mean(pnls), 2) if pnls else 0
        total_p= round(sum(pnls), 2)
        wr     = f"{wins/len(grp)*100:.1f}%"
        grade_fill = fill(C_WIN if grade in ("A+","A") else C_NEUTRAL if grade == "B" else C_LOSS)
        for ci, v in enumerate([grade, len(grp), wins, wr, avg_p, total_p], start=1):
            c = ws_fail.cell(row=row, column=ci, value=v)
            c.font      = body_font(bold=(grade in ("A+","A")))
            c.border    = border()
            c.alignment = center()
            c.fill      = grade_fill
        row += 1

    # Recommendations
    row += 2
    ws_fail.cell(row=row, column=1, value="ALGORITHM RECOMMENDATIONS").font = hdr_font(11, color="000000")
    ws_fail.cell(row=row, column=1).fill = fill(C_SUBHDR)
    row += 1
    recs = _generate_recommendations(results)
    for rec in recs:
        ws_fail.cell(row=row, column=1, value=rec).alignment = left(wrap=True)
        ws_fail.cell(row=row, column=1).font = body_font(9)
        ws_fail.merge_cells(f"A{row}:H{row}")
        ws_fail.row_dimensions[row].height = 25
        row += 1

    set_col_widths(ws_fail, [45, 10, 14, 50, 12, 12, 12, 12])

    # ══════════════════════════════════════════════════════════════════════
    # SHEET: NIFTY Levels (2-year analysis)
    # ══════════════════════════════════════════════════════════════════════
    ws_lv = wb.create_sheet("NIFTY Key Levels")
    ws_lv.sheet_view.showGridLines = False

    ws_lv.merge_cells("A1:N1")
    ws_lv["A1"].value     = "NIFTY — KEY S/R LEVELS  (2-Year Daily + Weekly Analysis)"
    ws_lv["A1"].font      = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws_lv["A1"].fill      = fill(C_HEADER)
    ws_lv["A1"].alignment = center()
    ws_lv.row_dimensions[1].height = 26

    # Summary stats
    for si, srow in enumerate(levels_summary.itertuples(), start=3):
        ws_lv.cell(row=si, column=1, value=srow.Metric).font = body_font(bold=True)
        ws_lv.cell(row=si, column=2, value=srow.Value)
        ws_lv.cell(row=si, column=1).border = border()
        ws_lv.cell(row=si, column=2).border = border()

    if not levels_df.empty:
        lvl_row = 3 + len(levels_summary) + 2
        lv_headers = list(levels_df.columns)
        write_header_row(ws_lv, lvl_row, lv_headers, col_fill=C_SUBHDR)
        for li, lrow in enumerate(levels_df.itertuples(index=False), start=lvl_row + 1):
            ltype = lrow[1] if len(lrow) > 1 else ""
            row_fill = fill(C_LEVEL_S if ltype == "SUPPORT" else C_LEVEL_R)
            for ci, v in enumerate(lrow, start=1):
                c = ws_lv.cell(row=li, column=ci, value=v)
                c.font      = body_font()
                c.border    = border()
                c.alignment = center()
                c.fill      = row_fill if ci <= 4 else fill("FFFFFF")

    set_col_widths(ws_lv, [10, 12, 12, 9, 10, 12, 12, 12, 12, 16, 16, 12, 10, 10])

    wb.save(out_path)
    logger.info(f"Excel saved: {out_path}")


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _failure_insight(reason: str) -> str:
    insights = {
        "Entered too far from level (APPROACHING zone, low probability)":
            "FIX: Only take trades in AT_LEVEL zone (dist_atr < 0.3). "
            "APPROACHING zone has ~30% lower win rate. Tighten proximity filter.",
        "Weak level (only 2 touches — insufficient confirmation)":
            "FIX: Raise MIN_LEVEL_TOUCHES to 3. Levels with 2 touches lack "
            "confirmation; often break rather than bounce.",
        "Low-strength level — below 15MIN timeframe":
            "FIX: Ignore 15MIN levels for swing trades. Only use DAILY/WEEKLY/HOURLY "
            "levels with strength > 0.4 for swing entries.",
        "SL hit — valid setup but adverse price action":
            "INSIGHT: Valid setup that failed due to market conditions. "
            "Consider wider SL (50pts) with proportionally smaller position size.",
        "Position held too long — no clear breakout in direction":
            "FIX: If trade hasn't moved 15pts in our favor within 3 bars, "
            "exit at breakeven. Time-based exits improve overall results.",
        "Trade not resolved within session":
            "FIX: Always set hard EOD close. For swing trades, roll over only "
            "if the level has held and trend is intact.",
    }
    return insights.get(reason,
        "Review trade context — check if broader market trend conflicted with setup.")


def _generate_recommendations(results) -> List[str]:
    """Generate algorithm improvement recommendations from backtest data."""
    all_trades = [t for r in results for t in r.trades]
    if not all_trades:
        return ["No trades found — check data availability and signal thresholds."]

    recs = []

    # Win rate analysis
    total = len(all_trades)
    wins  = sum(1 for t in all_trades if t.is_winner)
    wr    = wins / total * 100 if total else 0

    if wr < 40:
        recs.append(
            f"⚠️  Win rate {wr:.1f}% is below target (40%). "
            "Recommendation: Raise MIN_LEVEL_TOUCHES to 3, "
            "restrict to AT_LEVEL zone only (dist_atr < 0.25), "
            "and require DAILY or WEEKLY timeframe levels only."
        )
    elif wr >= 55:
        recs.append(
            f"✅  Win rate {wr:.1f}% is strong. "
            "Consider increasing position size on A+ setups by 25%."
        )

    # Setup grade analysis
    for grade in ["A+", "A", "B", "C"]:
        grp = [t for t in all_trades if t.setup_quality == grade]
        if len(grp) < 3:
            continue
        g_wr = sum(1 for t in grp if t.is_winner) / len(grp) * 100
        if grade == "C" and g_wr < 35:
            recs.append(
                f"🔴  C-grade setups win rate: {g_wr:.1f}%. "
                "Recommendation: SKIP all C-grade setups entirely. "
                "They drag down overall performance."
            )
        elif grade == "B" and g_wr < 40:
            recs.append(
                f"🟡  B-grade setups win rate: {g_wr:.1f}%. "
                "Recommendation: Only take B-grade at DAILY/WEEKLY levels with ≥4 touches."
            )

    # SL hit analysis
    sl_hits = [t for t in all_trades if t.exit_reason == "SL_HIT"]
    if sl_hits and total:
        sl_pct = len(sl_hits) / total * 100
        if sl_pct > 50:
            recs.append(
                f"🔴  {sl_pct:.1f}% of trades hit SL. "
                "Current SL (Daily: 40pts, Intraday: 15pts) may be too tight. "
                "Recommendation: Widen daily SL to 50pts but reduce position size by 20%, "
                "keeping max risk per trade constant."
            )

    # Max hold exits
    maxhold = [t for t in all_trades if t.exit_reason == "MAX_HOLD"]
    if maxhold:
        avg_pnl = np.mean([t.pnl_rupees for t in maxhold])
        recs.append(
            f"⏰  {len(maxhold)} trades hit max-hold limit. Avg P&L on these: ₹{avg_pnl:+,.0f}. "
            "{'Consider cutting max-hold from 7→5 days.' if avg_pnl < 0 else 'Max-hold performing OK.'}"
        )

    # Proximity zone analysis
    approaching = [t for t in all_trades if t.proximity_zone == "APPROACHING"]
    at_level    = [t for t in all_trades if t.proximity_zone == "AT_LEVEL"]
    if approaching and at_level:
        wr_app = sum(1 for t in approaching if t.is_winner) / len(approaching) * 100
        wr_atl = sum(1 for t in at_level if t.is_winner) / len(at_level) * 100
        recs.append(
            f"📍  AT_LEVEL win rate: {wr_atl:.1f}%  |  "
            f"APPROACHING win rate: {wr_app:.1f}%. "
            f"{'APPROACHING signals are significantly weaker — consider disabling them.' if wr_app < wr_atl - 10 else 'Both zones performing similarly.'}"
        )

    # Level timeframe analysis
    for tf in ["WEEKLY", "DAILY", "HOURLY", "15MIN"]:
        grp = [t for t in all_trades if t.level_tf == tf]
        if len(grp) < 3:
            continue
        g_wr = sum(1 for t in grp if t.is_winner) / len(grp) * 100
        recs.append(
            f"📊  {tf} levels: {len(grp)} trades, {g_wr:.1f}% win rate."
        )

    if not recs:
        recs.append("✅  Algorithm performing as expected. No major issues detected.")

    return recs


def print_console_summary(results):
    print()
    print("=" * 72)
    print(f"  {'PERIOD':<12} {'TRADES':>7} {'WIN%':>7} {'P&L (Rs)':>13} {'PF':>6} {'SHARPE':>7} {'MAX DD':>11}")
    print("-" * 72)
    for r in results:
        sign = "+" if r.total_pnl >= 0 else ""
        print(
            f"  {r.label:<12} {r.total_trades:>7} "
            f"{r.win_rate:>6.1f}% "
            f"Rs{sign}{r.total_pnl:>10,.0f} "
            f"{r.profit_factor:>6.2f} "
            f"{r.sharpe_ratio:>7.2f} "
            f"Rs{r.max_drawdown:>9,.0f}"
        )
    print("=" * 72)


if __name__ == "__main__":
    results, path = run_backtest()
